#!/usr/bin/env python3
"""Aplicação GeRot focada em visibilidade de dashboards e agenda diária."""

from __future__ import annotations

from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    session,
    jsonify,
    request,
    flash,
    send_from_directory,
    g,
)
from flask_cors import CORS
from flask_compress import Compress
from flask_restful import Api, Resource
from dotenv import load_dotenv
import os
import time

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

import secrets
import re
from pathlib import Path
import bcrypt
import psycopg2
import psycopg2.extras
import psycopg2.errors
from datetime import datetime, date, timedelta
from functools import wraps
from typing import Dict, List, Tuple
import requests
import mimetypes
from io import BytesIO
from werkzeug.utils import secure_filename
from psycopg2 import pool
import openai
import google.generativeai as genai
import google.generativeai as genai

from openpyxl import load_workbook

from utils.planner_client import PlannerClient, PlannerIntegrationError


app = Flask(__name__)
CORS(app)
Compress(app)  # Habilita compressão Gzip
api = Api(app)


@app.after_request
def add_header(response):
    if 'Cache-Control' not in response.headers:
        # Cache para arquivos estáticos (static/...)
        if request.path.startswith('/static/'):
            response.headers['Cache-Control'] = 'public, max-age=31536000'
        # Não cachear API ou HTML dinâmico para evitar dados velhos
        else:
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

# --------------------------------------------------------------------------- #
# Configuração base
# --------------------------------------------------------------------------- #
app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY", "gerot-production-2025-super-secret"
)
app.config["DEBUG"] = os.getenv("FLASK_DEBUG", "false").lower() == "true"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)  # Sessões persistem por 7 dias

DATABASE_URL = (
    os.getenv("DATABASE_URL")
    or os.getenv("DIRECT_URL")
    or os.getenv("SUPABASE_DB_URL")
)

if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL não configurada. Defina a string de conexão do Supabase."
    )

app.config["DATABASE_URL"] = DATABASE_URL

# Configuração OpenAI
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY
else:
    print("⚠️ AVISO: OPENAI_API_KEY não configurada. O Chat IA não funcionará corretamente.")

# Configuração Google Gemini
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
if GOOGLE_API_KEY:
    genai.configure(api_key=GOOGLE_API_KEY)
else:
    print("⚠️ AVISO: GOOGLE_API_KEY não configurada. Fallback para Gemini inativo.")

# Limpar URL para o Pool (remover pgbouncer e outros params incompatíveis)
pool_dsn = DATABASE_URL
if "?" in pool_dsn:
    url_parts = pool_dsn.split("?")
    base_url = url_parts[0]
    if len(url_parts) > 1:
        query_params = url_parts[1]
        # Remove parâmetros pgbouncer
        params = [p for p in query_params.split("&") if not p.startswith("pgbouncer=")]
        if params:
            pool_dsn = f"{base_url}?{'&'.join(params)}"
        else:
            pool_dsn = base_url

# Configuração do Pool de Conexões
# Min: 1, Max: 60 conexões simultâneas
try:
    db_pool = psycopg2.pool.ThreadedConnectionPool(
        minconn=1,
        maxconn=60,
        dsn=pool_dsn,  # Usar URL limpa
        cursor_factory=psycopg2.extras.RealDictCursor,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=5
    )
    print("✅ Pool de conexões criado com sucesso (1-60 conexões)")
except Exception as e:
    print(f"❌ Erro fatal ao criar pool de conexões: {e}")
    # Fallback ou exit? Vamos deixar passar e tentar reconectar no get_db se falhar
    db_pool = None


BASE_DIR = Path(__file__).resolve().parent
PLANILHA_USUARIOS = BASE_DIR / "dados.xlsx"
ADMIN_CARGOS = {"CONSULTOR", "COORDENADOR", "DIRETOR"}
ALLOWED_AVATAR_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}
MAX_AVATAR_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB

# Configuração para usar o novo tema Tailwind (True) ou o tema antigo (False)
USE_TAILWIND_THEME = os.getenv("USE_TAILWIND_THEME", "true").lower() == "true"

DEFAULT_DASHBOARDS = [
    {
        "slug": "comercial_sc",
        "title": "Comercial SC",
        "description": "Relatório Vendas PortoEx - unidade SC.",
        "category": "Comercial",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNDAwZTA5YjgtZWVlMC00MzQ2LWJmYmQtYTZiZDVlMDhlZTEyIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 1,
    },
    {
        "slug": "comercial_sp",
        "title": "Comercial SP",
        "description": "Relatório Vendas PortoEx - filial São Paulo.",
        "category": "Comercial",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYjMyZTc5MzktNGFhYi00ZjE1LWFjMDctYjY2ODM4NTlhMWRmIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 2,
    },
    {
        "slug": "controladoria_target",
        "title": "Controladoria e Target Operação",
        "description": "Visão consolidada da controladoria e operação.",
        "category": "Controladoria",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiMzhjODY4OGYtY2UxMy00ZjkyLTkzNDEtOTcxZWIzNDY2ZGJlIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 3,
    },
    {
        "slug": "cotacao_sc",
        "title": "Cotação SC",
        "description": "Painel de cotações da filial Santa Catarina.",
        "category": "Cotação",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNDcwMmM0NGUtOGJkZC00NmIyLTk3M2QtOTNjOWEzMDY5MjkwIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 4,
    },
    {
        "slug": "cotacao_sp",
        "title": "Cotação SP",
        "description": "Painel de cotações da filial São Paulo.",
        "category": "Cotação",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYmNkMWQxMjUtZjExNC00ZGE5LWIxYTEtYzlmNzI3M2I3Mjg1IiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 5,
    },
    {
        "slug": "gr_leandro",
        "title": "Gestão Regional - Leandro",
        "description": "Indicadores táticos da regional do Leandro.",
        "category": "Operações",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYjI4YTIzMDEtZmRmOC00N2Y3LTkzNmQtMGEwYzE1N2VhMDViIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 6,
    },
    {
        "slug": "financeiro_fluxo",
        "title": "Financeiro - Fluxo de Caixa",
        "description": "Inadimplência e fluxo de caixa diário.",
        "category": "Financeiro",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNzFlZWVkZjUtNTdiYy00ZjJiLTk3OTEtNzhiYzFhMzk3MmY3IiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 7,
    },
]

PLANNER_CONFIG = {
    "tenant_id": os.getenv("MS_TENANT_ID"),
    "client_id": os.getenv("MS_CLIENT_ID"),
    "client_secret": os.getenv("MS_CLIENT_SECRET"),
    "plan_id": os.getenv("MS_PLANNER_PLAN_ID"),
    "bucket_id": os.getenv("MS_PLANNER_BUCKET_ID"),
}

planner_client = PlannerClient(**PLANNER_CONFIG)


# --------------------------------------------------------------------------- #
# Decorators e utilidades
# --------------------------------------------------------------------------- #
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Por favor, faça login.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Acesso restrito aos administradores.", "error")
            return redirect(url_for("team_dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def is_admin_session() -> bool:
    return session.get("role") == "admin"


def get_template(template_base_name: str) -> str:
    """Retorna o template correto baseado na configuração USE_TAILWIND_THEME."""
    if USE_TAILWIND_THEME:
        # Verifica se existe uma versão Tailwind do template
        tailwind_template = f"{template_base_name.replace('.html', '')}_tailwind.html"
        template_path = Path(BASE_DIR) / "templates" / tailwind_template
        if template_path.exists():
            return tailwind_template
    return template_base_name


def _as_bytes(value):
    if value is None:
        return None
    if isinstance(value, memoryview):
        return value.tobytes()
    if isinstance(value, str):
        return value.encode("utf-8")
    return value


def _sanitize_optional(value: str | None):
    if value is None:
        return None
    value = value.strip()
    return value or None


def is_allowed_avatar_file(filename: str) -> bool:
    return bool(filename and "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_AVATAR_EXTENSIONS)


def refresh_session_user_cache():
    """Recarrega dados básicos do usuário na sessão após atualização."""
    if "user_id" not in session:
        return None
    updated = get_user_by_id(session["user_id"])
    if updated:
        session["username"] = updated.get("username", session.get("username"))
        session["nome_completo"] = updated.get("nome_completo", session.get("nome_completo"))
        session["role"] = updated.get("role", session.get("role"))
        if updated.get("departamento") is not None:
            session["departamento"] = updated.get("departamento")
        if updated.get("avatar_url"):
            session["avatar_url"] = updated["avatar_url"]
    return updated


@app.before_request
def update_last_seen():
    if session.get('user_id'):
        # Ignora rotas estáticas
        if request.path.startswith('/static') or request.endpoint == 'static':
            return

        try:
            conn = get_db()
            cursor = conn.cursor()
            # Usar request.path para mostrar a URL real
            current_page = request.path
            
            cursor.execute("""
                UPDATE users_new 
                SET last_seen_at = CURRENT_TIMESTAMP, current_page = %s 
                WHERE id = %s
            """, (current_page, session['user_id']))
            conn.commit()
            conn.close()
        except Exception as e:
            # Logar erro para debug
            app.logger.error(f"Erro ao atualizar last_seen: {e}")


@app.route("/admin/live-users")
@login_required
@admin_required
def admin_live_users():
    conn = get_db()
    cursor = conn.cursor()
    # Buscar usuários ativos nos últimos 5 minutos
    cursor.execute("""
        SELECT id, nome_completo, username, role, last_seen_at, current_page,
        EXTRACT(EPOCH FROM (NOW() - last_seen_at)) as seconds_ago
        FROM users_new
        WHERE last_seen_at > NOW() - INTERVAL '5 minutes'
        ORDER BY last_seen_at DESC
    """)
    active_users = cursor.fetchall()
    
    conn.close()
    return render_template(get_template("admin_live_users.html"), users=active_users)


@app.route("/admin/debug-time")
@login_required
@admin_required
def debug_time():
    """Rota temporária para diagnosticar problemas de time/fuso"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Verifica hora do DB
    cursor.execute("SELECT NOW() as db_time, CURRENT_TIMESTAMP as db_timestamp, NOW() - INTERVAL '5 minutes' as cutoff")
    times = cursor.fetchone()
    
    # Verifica últimos updates reais
    cursor.execute("""
        SELECT username, last_seen_at, current_page,
        NOW() - last_seen_at as time_diff
        FROM users_new
        WHERE last_seen_at IS NOT NULL
        ORDER BY last_seen_at DESC
        LIMIT 5
    """)
    recent_users = cursor.fetchall()
    conn.close()
    
    return jsonify({
        "server_time_utc": str(datetime.utcnow()),
        "db_time": str(times['db_time']),
        "cutoff_5min": str(times['cutoff']),
        "recent_users": [
            {
                "username": r['username'],
                "last_seen": str(r['last_seen_at']),
                "page": r['current_page'],
                "diff": str(r['time_diff'])
            } for r in recent_users
        ]
    })


# Classe Wrapper para interceptar o close() sem modificar o objeto C do psycopg2
class ConnectionWrapper:
    def __init__(self, conn, from_pool=True):
        self._conn = conn
        self._from_pool = from_pool
    
    def close(self):
        # Ignora chamadas explícitas de close() no código legado
        pass
    
    def real_close(self):
        # Fecha de verdade (usado para conexões fora do pool)
        self._conn.close()
        
    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_db():
    """Obtém conexão do pool ou cria nova se necessário."""
    if 'db_wrapper' not in g:
        # Tentar pegar do Pool
        if db_pool:
            try:
                conn = db_pool.getconn()
                g.db_wrapper = ConnectionWrapper(conn, from_pool=True)
                return g.db_wrapper
            except Exception as e:
                app.logger.error(f"[DB] Erro ao pegar do pool: {e}")
        
        # Fallback: conexão direta (sem pool)
        database_url = app.config["DATABASE_URL"]
        if "?" in database_url:
            url_parts = database_url.split("?")
            base_url = url_parts[0]
            if len(url_parts) > 1:
                query_params = url_parts[1]
                params = [p for p in query_params.split("&") if not p.startswith("pgbouncer=")]
                database_url = f"{base_url}?{'&'.join(params)}" if params else base_url
        
        max_retries = 3
        last_error = None
        for attempt in range(max_retries):
            try:
                conn = psycopg2.connect(
                    database_url,
                    cursor_factory=psycopg2.extras.RealDictCursor,
                    keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5
                )
                g.db_wrapper = ConnectionWrapper(conn, from_pool=False)
                return g.db_wrapper
            except psycopg2.OperationalError as e:
                last_error = e
                time.sleep(1 * (attempt + 1))
        raise last_error

    return g.db_wrapper


@app.teardown_appcontext
def close_db(error):
    """Devolve a conexão ao pool ou fecha ao final da requisição."""
    wrapper = g.pop('db_wrapper', None)
    if wrapper is not None:
        conn = wrapper._conn
        if wrapper._from_pool and db_pool:
            try:
                db_pool.putconn(conn)
            except Exception as e:
                app.logger.error(f"[DB] Erro ao devolver ao pool: {e}")
        else:
            # Se não veio do pool, fecha de verdade
            try:
                conn.close()
            except:
                pass


def ensure_schema() -> None:
    conn = get_db()
    cursor = conn.cursor()
    
    # Adquirir lock exclusivo para garantir que apenas um worker execute a migração
    cursor.execute("SELECT pg_advisory_xact_lock(12345)")

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users_new (
            id BIGSERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password BYTEA NOT NULL,
            nome_completo TEXT NOT NULL,
            cargo_original TEXT,
            departamento TEXT,
            role TEXT NOT NULL DEFAULT 'usuario',
            email TEXT,
            nome_usuario TEXT,
            unidade TEXT,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            first_login BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ,
            last_login TIMESTAMPTZ
        );
        """
    )
    
    # Adicionar coluna nome_usuario se não existir (migration)
    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = 'users_new' AND column_name = 'nome_usuario'
            ) THEN
                ALTER TABLE users_new ADD COLUMN nome_usuario TEXT;
                CREATE UNIQUE INDEX IF NOT EXISTS users_new_nome_usuario_unique
                    ON users_new (LOWER(nome_usuario)) WHERE nome_usuario IS NOT NULL;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = 'users_new' AND column_name = 'avatar_url'
            ) THEN
                ALTER TABLE users_new ADD COLUMN avatar_url TEXT;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS users_new_email_unique
            ON users_new (LOWER(email));
        """
    )
    
    # Adicionar colunas para rastreamento de atividade em tempo real
    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = 'users_new' AND column_name = 'last_seen_at') THEN
                ALTER TABLE users_new ADD COLUMN last_seen_at TIMESTAMPTZ;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = 'users_new' AND column_name = 'current_page') THEN
                ALTER TABLE users_new ADD COLUMN current_page TEXT;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS users_new_username_lower_unique
            ON users_new (LOWER(username));
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboards (
            id BIGSERIAL PRIMARY KEY,
            slug TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT,
            embed_url TEXT NOT NULL,
            display_order INTEGER NOT NULL DEFAULT 0,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS user_dashboards (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL REFERENCES users_new(id) ON DELETE CASCADE,
            dashboard_id BIGINT NOT NULL REFERENCES dashboards(id) ON DELETE CASCADE,
            created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            UNIQUE (user_id, dashboard_id)
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS planner_sync_logs (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            user_name TEXT,
            dashboard_count INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL,
            message TEXT,
            task_id TEXT,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    conn.commit()
    conn.close()
    
    # Criar usuário admin anaissiabraao se não existir
    create_admin_user()


def create_admin_user() -> None:
    """Cria o usuário admin anaissiabraao com todos os privilégios"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Verifica se o usuário já existe
        cursor.execute(
            """
            SELECT id FROM users_new 
            WHERE LOWER(username) = LOWER('anaissiabraao@gmail.com')
               OR LOWER(nome_usuario) = LOWER('anaissiabraao')
            """
        )
        existing = cursor.fetchone()
        
        if not existing:
            # Cria senha padrão
            temp_password = "admin123"  # Senha temporária, deve ser alterada no primeiro acesso
            password_hash = bcrypt.hashpw(
                temp_password.encode("utf-8"), bcrypt.gensalt()
            )
            
            cursor.execute(
                """
                INSERT INTO users_new (
                    username, password, nome_completo, cargo_original,
                    departamento, role, email, nome_usuario, is_active, first_login
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (username) DO UPDATE SET
                    nome_usuario = EXCLUDED.nome_usuario,
                    role = EXCLUDED.role,
                    is_active = EXCLUDED.is_active
                """,
                (
                    "anaissiabraao@gmail.com",
                    psycopg2.Binary(password_hash),
                    "ABRAAO DE OLIVEIRA COSTA ANAISSI",
                    "DIRETOR",
                    "ADMINISTRATIVO",
                    "admin",
                    "anaissiabraao@portoex.com.br",
                    "anaissiabraao",
                    True,
                    True,  # Primeiro acesso
                ),
            )
            conn.commit()
            app.logger.info("[ADMIN] Usuário admin anaissiabraao criado com sucesso")
        else:
            # Atualiza o usuário existente para garantir que é admin (sem resetar senha)
            cursor.execute(
                """
                UPDATE users_new
                SET nome_usuario = 'anaissiabraao',
                    role = 'admin',
                    is_active = true
                WHERE id = %s
                """,
                (existing["id"],),
            )
            conn.commit()
            app.logger.info("[ADMIN] Usuário admin anaissiabraao atualizado (sem resetar senha)")
        
        conn.close()
    except Exception as exc:
        app.logger.error(f"[ADMIN] Erro ao criar usuário admin: {exc}")


def ensure_agent_tables():
    """Garante que as tabelas do agente existam via SQL direto."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Verificar se tabelas já existem para evitar processamento desnecessário
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'agent_rpa_types'
            )
        """)
        if cursor.fetchone()['exists']:
            conn.close()
            return

        app.logger.info("[AGENT] Criando tabelas do Agente IA...")
        
        # Criação das tabelas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_rpa_types (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                icon TEXT DEFAULT 'fa-cogs',
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_rpas (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                rpa_type_id BIGINT REFERENCES agent_rpa_types(id) ON DELETE SET NULL,
                priority TEXT NOT NULL DEFAULT 'medium',
                frequency TEXT DEFAULT 'once',
                parameters JSONB,
                status TEXT NOT NULL DEFAULT 'pending',
                result JSONB,
                error_message TEXT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                executed_at TIMESTAMPTZ,
                completed_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_rpas_status ON agent_rpas(status);
            CREATE INDEX IF NOT EXISTS idx_agent_rpas_created_by ON agent_rpas(created_by);

            CREATE TABLE IF NOT EXISTS agent_data_sources (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                source_type TEXT NOT NULL DEFAULT 'database',
                connection_config JSONB,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_dashboard_requests (
                id BIGSERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT NOT NULL DEFAULT 'Outros',
                data_source_id BIGINT REFERENCES agent_data_sources(id) ON DELETE SET NULL,
                chart_types TEXT[],
                filters JSONB,
                status TEXT NOT NULL DEFAULT 'pending',
                result_url TEXT,
                result_data JSONB,
                error_message TEXT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                processed_at TIMESTAMPTZ,
                completed_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_status ON agent_dashboard_requests(status);
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_created_by ON agent_dashboard_requests(created_by);

            CREATE TABLE IF NOT EXISTS agent_settings (
                id BIGSERIAL PRIMARY KEY,
                setting_key TEXT NOT NULL UNIQUE,
                setting_value JSONB,
                description TEXT,
                updated_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_logs (
                id BIGSERIAL PRIMARY KEY,
                action_type TEXT NOT NULL,
                entity_type TEXT,
                entity_id BIGINT,
                user_id BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                details JSONB,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_logs_action_type ON agent_logs(action_type);
            
            CREATE TABLE IF NOT EXISTS agent_dashboard_templates (
                id BIGSERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT NOT NULL DEFAULT 'Outros',
                data_source_id BIGINT REFERENCES agent_data_sources(id) ON DELETE SET NULL,
                query_config JSONB,
                layout_config JSONB,
                charts_config JSONB,
                filters_config JSONB,
                theme_config JSONB,
                is_published BOOLEAN DEFAULT false,
                is_public BOOLEAN DEFAULT false,
                thumbnail_url TEXT,
                linked_dashboard_id BIGINT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_templates_created_by ON agent_dashboard_templates(created_by);
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_templates_published ON agent_dashboard_templates(is_published);
        """)

        # Inserção de dados iniciais
        cursor.execute("""
            INSERT INTO agent_rpa_types (name, description, icon) VALUES
            ('Extração de Dados', 'Extrai dados de sistemas externos (ERP, planilhas, APIs)', 'fa-download'),
            ('Processamento de Arquivos', 'Processa e transforma arquivos (PDF, Excel, CSV)', 'fa-file-alt'),
            ('Integração de Sistemas', 'Sincroniza dados entre sistemas diferentes', 'fa-sync'),
            ('Envio de Relatórios', 'Gera e envia relatórios automaticamente', 'fa-paper-plane'),
            ('Monitoramento', 'Monitora sistemas e envia alertas', 'fa-bell'),
            ('Backup de Dados', 'Realiza backup automático de dados', 'fa-database'),
            ('Web Scraping', 'Coleta dados de websites', 'fa-globe'),
            ('Automação de E-mail', 'Processa e responde e-mails automaticamente', 'fa-envelope')
            ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description;

            INSERT INTO agent_data_sources (name, description, source_type) VALUES
            ('Banco de Dados GeRot', 'Dados internos do sistema GeRot', 'database'),
            ('Power BI', 'Dados dos dashboards Power BI', 'api'),
            ('Planilhas Excel', 'Dados de planilhas compartilhadas', 'file'),
            ('ERP PortoEx', 'Sistema ERP da empresa', 'api'),
            ('API Externa', 'Dados de APIs de terceiros', 'api')
            ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description;
            
            INSERT INTO agent_settings (setting_key, setting_value, description) VALUES
            ('rpa_enabled', '{"enabled": true}', 'Habilita/desabilita funcionalidades de RPA'),
            ('dashboard_gen_enabled', '{"enabled": true}', 'Habilita/desabilita geração de dashboards'),
            ('max_concurrent_rpas', '{"value": 5}', 'Número máximo de RPAs executando simultaneamente')
            ON CONFLICT (setting_key) DO NOTHING;
        """)
        
        conn.commit()
        conn.close()
        app.logger.info("[AGENT] Tabelas e dados iniciais criados com sucesso.")
        
    except Exception as e:
        app.logger.error(f"[AGENT] Erro ao criar tabelas: {e}")


def seed_dashboards() -> None:
    conn = get_db()
    cursor = conn.cursor()

    for dash in DEFAULT_DASHBOARDS:
        cursor.execute(
            """
            INSERT INTO dashboards (slug, title, description, category, embed_url, display_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(slug) DO UPDATE SET
                title=excluded.title,
                description=excluded.description,
                category=excluded.category,
                embed_url=excluded.embed_url,
                display_order=excluded.display_order,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                dash["slug"],
                dash["title"],
                dash["description"],
                dash["category"],
                dash["embed_url"],
                dash["display_order"],
                True,  # is_active como boolean
            ),
        )

    conn.commit()
    conn.close()


def normalize_roles() -> None:
    """Normaliza roles dos usuários com retry para evitar deadlocks"""
    max_retries = 3
    conn = None
    for attempt in range(max_retries):
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            # Otimização: Só atualiza se o valor for diferente
            # Isso evita locks desnecessários em linhas que já estão corretas
            
            # 1. Normalizar admin_master -> admin
            cursor.execute(
                """
                UPDATE users_new
                SET role = 'admin'
                WHERE role = 'admin_master'
                """
            )
            
            # 2. Definir 'usuario' para quem não é admin e não é usuario (ex: null ou outros)
            cursor.execute(
                """
                UPDATE users_new
                SET role = 'usuario'
                WHERE role IS NULL OR (role != 'admin' AND role != 'usuario')
                """
            )
            
            conn.commit()
            conn.close()
            return
        except psycopg2.errors.DeadlockDetected:
            if attempt < max_retries - 1:
                import time
                time.sleep(0.1 * (attempt + 1))  # Backoff exponencial
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                    conn.close()
                continue
            else:
                app.logger.warning("[normalize_roles] Aviso: deadlock detected após múltiplas tentativas", exc_info=True)
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                    conn.close()
                return
        except Exception as exc:
            app.logger.warning(f"[normalize_roles] Aviso: {exc}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
                conn.close()
            return
        

def _determine_role_from_cargo(cargo: str | None) -> str:
    cargo_normalizado = (cargo or "").strip().upper()
    return "admin" if cargo_normalizado in ADMIN_CARGOS else "usuario"


def import_users_from_excel() -> None:
    if not PLANILHA_USUARIOS.exists():
        app.logger.warning(
            "[IMPORTACAO] Arquivo dados.xlsx não encontrado em %s",
            PLANILHA_USUARIOS,
        )
        return

    try:
        workbook = load_workbook(
            filename=str(PLANILHA_USUARIOS), read_only=True, data_only=True
        )
    except Exception as exc:
        app.logger.error(
            "[IMPORTACAO] Falha ao abrir planilha de usuários: %s", exc
        )
        return

    conn = None
    inserted = updated = skipped = 0

    try:
        conn = get_db()
        cursor = conn.cursor()
        for row in workbook.active.iter_rows(min_row=2, values_only=True):
            if not row:
                skipped += 1
                continue

            values = list(row)
            while len(values) < 5:
                values.append(None)

            nome, email, cargo, unidade, departamento = values[:5]
            email = (email or "").strip()
            if not email:
                skipped += 1
                continue

            nome = (nome or "").strip() or email
            cargo = (cargo or "").strip()
            unidade = (unidade or "").strip()
            departamento = (departamento or "").strip()
            role = _determine_role_from_cargo(cargo)
            username = email.lower()

            cursor.execute(
                "SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s)", (email,)
            )
            existing = cursor.fetchone()

            if existing:
                # Retry em caso de deadlock
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        cursor.execute(
                            """
                            UPDATE users_new
                            SET nome_completo = %s,
                                cargo_original = %s,
                                departamento = %s,
                                unidade = %s,
                                role = %s,
                                updated_at = NOW()
                            WHERE id = %s
                            """,
                            (
                                nome,
                                cargo or None,
                                departamento or None,
                                unidade or None,
                                role,
                                existing["id"],
                            ),
                        )
                        updated += 1
                        break
                    except psycopg2.errors.DeadlockDetected:
                        if attempt < max_retries - 1:
                            import time
                            import random
                            time.sleep(0.1 * (attempt + 1) + random.uniform(0, 0.1))  # Backoff exponencial com jitter
                            conn.rollback()
                            continue
                        else:
                            app.logger.warning(
                                f"[IMPORTACAO] Deadlock após múltiplas tentativas para {email}, pulando usuário",
                                exc_info=True,
                            )
                            skipped += 1
                            break  # Pula este usuário e continua
            else:
                temp_password = secrets.token_urlsafe(16)
                password_hash = bcrypt.hashpw(
                    temp_password.encode("utf-8"), bcrypt.gensalt()
                )
                # Retry em caso de deadlock ou unique violation
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        cursor.execute(
                            """
                            INSERT INTO users_new (
                                username,
                                password,
                                nome_completo,
                                cargo_original,
                                departamento,
                                role,
                                email,
                                unidade,
                                first_login,
                                nome_usuario
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE, NULL)
                            ON CONFLICT (username) DO UPDATE SET
                                nome_completo = EXCLUDED.nome_completo,
                                cargo_original = EXCLUDED.cargo_original,
                                departamento = EXCLUDED.departamento,
                                unidade = EXCLUDED.unidade,
                                role = EXCLUDED.role,
                                email = EXCLUDED.email,
                                updated_at = NOW()
                            """,
                            (
                                username,
                                psycopg2.Binary(password_hash),
                                nome,
                                cargo or None,
                                departamento or None,
                                role,
                                email,
                                unidade or None,
                            ),
                        )
                        # ON CONFLICT sempre retorna rowcount > 0, precisa verificar se foi insert ou update
                        if cursor.rowcount > 0:
                            # Verifica se foi realmente insert (created_at == updated_at) ou update
                            cursor.execute("SELECT created_at, updated_at FROM users_new WHERE username = %s", (username,))
                            user_check = cursor.fetchone()
                            if user_check and user_check.get("created_at") == user_check.get("updated_at"):
                                inserted += 1
                            else:
                                updated += 1
                        break
                    except (psycopg2.errors.DeadlockDetected, psycopg2.errors.UniqueViolation) as e:
                        if attempt < max_retries - 1:
                            import time
                            import random
                            time.sleep(0.1 * (attempt + 1) + random.uniform(0, 0.1))  # Backoff exponencial com jitter
                            conn.rollback()
                            # Tenta novamente como UPDATE se for unique violation
                            if isinstance(e, psycopg2.errors.UniqueViolation):
                                cursor.execute("SELECT id FROM users_new WHERE username = %s", (username,))
                                existing = cursor.fetchone()
                                if existing:
                                    # Tenta fazer UPDATE em vez de INSERT
                                    try:
                                        cursor.execute(
                                            """
                                            UPDATE users_new
                                            SET nome_completo = %s,
                                                cargo_original = %s,
                                                departamento = %s,
                                                unidade = %s,
                                                role = %s,
                                                email = %s,
                                                updated_at = NOW()
                                            WHERE id = %s
                                            """,
                                            (
                                                nome,
                                                cargo or None,
                                                departamento or None,
                                                unidade or None,
                                                role,
                                                email,
                                                existing["id"],
                                            ),
                                        )
                                        updated += 1
                                        break
                                    except psycopg2.errors.DeadlockDetected:
                                        continue
                            continue
                        else:
                            app.logger.warning(
                                f"[IMPORTACAO] Erro após múltiplas tentativas para {email}, pulando usuário",
                                exc_info=True,
                            )
                            skipped += 1
                            break  # Pula este usuário e continua

        conn.commit()
        app.logger.info(
            "[IMPORTACAO] Usuários sincronizados. Inseridos=%s | Atualizados=%s | Ignorados=%s",
            inserted,
            updated,
            skipped,
        )
    except Exception as exc:
        if conn:
            conn.rollback()
        app.logger.exception(
            "[IMPORTACAO] Erro ao sincronizar usuários da planilha: %s", exc
        )
    finally:
        if conn:
            conn.close()
        workbook.close()


with app.app_context():
    try:
        ensure_schema()
        # ensure_agent_tables() # Pode não estar definida neste arquivo, comentar se der erro
        seed_dashboards()
        normalize_roles()
        # import_users_from_excel() # Pode ser pesado para rodar em todo restart
    except Exception as e:
        app.logger.error(f"Erro na inicialização do banco: {e}")


# --------------------------------------------------------------------------- #
# Funções auxiliares de dados
# --------------------------------------------------------------------------- #
def fetch_dashboards(active_only: bool = True) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    query = "SELECT * FROM dashboards"
    if active_only:
        query += " WHERE is_active = true"
    query += " ORDER BY display_order, title"
    cursor.execute(query)
    dashboards = cursor.fetchall()
    conn.close()
    return dashboards


def fetch_users() -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, nome_completo, username, email, role
        FROM users_new
        WHERE is_active = true
        ORDER BY nome_completo
        """
    )
    users = cursor.fetchall()
    conn.close()
    return users


def get_user_dashboard_map() -> Dict[int, Dict[str, List]]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT ud.user_id, d.id as dashboard_id, d.title, d.category
        FROM user_dashboards ud
        JOIN dashboards d ON d.id = ud.dashboard_id
        WHERE d.is_active = true
        ORDER BY d.display_order, d.title
        """
    )
    data: Dict[int, Dict[str, List]] = {}
    for row in cursor.fetchall():
        entry = data.setdefault(row["user_id"], {"ids": set(), "items": []})
        entry["ids"].add(row["dashboard_id"])
        entry["items"].append({"title": row["title"], "category": row["category"]})

    conn.close()

    for entry in data.values():
        entry["items"].sort(key=lambda i: i["title"])
        entry["ids"] = list(entry["ids"])

    return data


def save_user_dashboards(user_id: int, dashboard_ids: List[int], actor_id: int) -> None:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_dashboards WHERE user_id = %s", (user_id,))
    if dashboard_ids:
        cursor.executemany(
            """
            INSERT INTO user_dashboards (user_id, dashboard_id, created_by)
            VALUES (%s, %s, %s)
            """,
            [(user_id, dash_id, actor_id) for dash_id in dashboard_ids],
        )
    conn.commit()
    conn.close()


def log_planner_sync(
    user_id: int,
    user_name: str,
    dashboard_count: int,
    status: str,
    message: str,
    task_id: str | None = None,
) -> None:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO planner_sync_logs (user_id, user_name, dashboard_count, status, message, task_id)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (user_id, user_name, dashboard_count, status, message, task_id),
    )
    conn.commit()
    conn.close()


def get_recent_planner_logs(limit: int = 8) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT user_name, dashboard_count, status, message, task_id, created_at
        FROM planner_sync_logs
        ORDER BY created_at DESC
        LIMIT %s
        """,
        (limit,),
    )
    logs = cursor.fetchall()
    conn.close()
    return logs


def sync_dashboards_to_planner() -> Tuple[int, List[str]]:
    if not planner_client.is_configured:
        raise PlannerIntegrationError(
            "Configure MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET, "
            "MS_PLANNER_PLAN_ID e MS_PLANNER_BUCKET_ID para usar esta função."
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT u.id as user_id, u.nome_completo, u.email,
               d.title, d.category, d.embed_url
        FROM user_dashboards ud
        JOIN users_new u ON u.id = ud.user_id
        JOIN dashboards d ON d.id = ud.dashboard_id
        WHERE u.is_active = true AND d.is_active = true
        ORDER BY u.nome_completo, d.display_order, d.title
        """
    )

    assignments: Dict[int, Dict[str, List[Dict[str, str]]]] = {}
    for row in cursor.fetchall():
        user_entry = assignments.setdefault(
            row["user_id"],
            {"name": row["nome_completo"], "email": row["email"], "dashboards": []},
        )
        user_entry["dashboards"].append(
            {
                "title": row["title"],
                "category": row["category"],
                "url": row["embed_url"],
            }
        )

    conn.close()

    if not assignments:
        raise PlannerIntegrationError("Nenhum usuário possui dashboards atribuídos.")

    successes = 0
    errors: List[str] = []
    today = date.today().strftime("%d/%m/%Y")
    start_time = datetime.utcnow().replace(hour=11, minute=0, second=0, microsecond=0)
    due_time = start_time + timedelta(hours=6)

    for user_id, payload in assignments.items():
        dashboards = payload["dashboards"]
        if not dashboards:
            continue

        title = f"Agenda de dashboards - {payload['name']} ({today})"
        description_lines = [
            f"Agenda automática gerada em {datetime.now().strftime('%d/%m/%Y %H:%M')}.",
            "",
            "Dashboards liberados para hoje:",
        ]
        for idx, dash in enumerate(dashboards, start=1):
            description_lines.append(
                f"{idx}. {dash['title']} ({dash['category']}) - {dash['url']}"
            )

        description = "\n".join(description_lines)

        try:
            task = planner_client.create_dashboard_task(
                title=title,
                description=description,
                start_time=start_time,
                due_time=due_time,
            )
            log_planner_sync(
                user_id=user_id,
                user_name=payload["name"],
                dashboard_count=len(dashboards),
                status="success",
                message="Tarefa criada e agenda enviada.",
                task_id=task.get("id"),
            )
            successes += 1
        except PlannerIntegrationError as exc:
            errors.append(f"{payload['name']}: {exc}")
            log_planner_sync(
                user_id=user_id,
                user_name=payload["name"],
                dashboard_count=len(dashboards),
                status="error",
                message=str(exc),
            )

    return successes, errors


# --------------------------------------------------------------------------- #
# Funções de autenticação
# --------------------------------------------------------------------------- #
def authenticate_user(identifier: str, password: str):
    try:
        conn = get_db()
        cursor = conn.cursor()
        # Aceita email OU nome_usuario OU username
        cursor.execute(
            """
            SELECT id, username, password, nome_completo, cargo_original,
                   departamento, role, email, nome_usuario, first_login
            FROM users_new
            WHERE (LOWER(username) = LOWER(%s) 
                   OR LOWER(email) = LOWER(%s)
                   OR (nome_usuario IS NOT NULL AND LOWER(nome_usuario) = LOWER(%s)))
              AND is_active = true
            """,
            (identifier, identifier, identifier),
        )
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            app.logger.debug(f"[AUTH] Usuário não encontrado: {identifier}")
            return None
        
        # Verifica senha
        password_valid = bcrypt.checkpw(password.encode("utf-8"), _as_bytes(user["password"]))
        if not password_valid:
            app.logger.debug(f"[AUTH] Senha incorreta para: {identifier}")
            return None
        
        app.logger.info(f"[AUTH] Login bem-sucedido: {identifier} (ID: {user['id']}, first_login: {user['first_login']})")
        role = "admin" if user["role"] == "admin" else "usuario"
        return {
            "id": user["id"],
            "username": user["username"],
            "nome_completo": user["nome_completo"],
            "cargo_original": user["cargo_original"],
            "departamento": user["departamento"],
            "role": role,
            "email": user["email"],
            "nome_usuario": user.get("nome_usuario"),
            "first_login": user["first_login"],
        }
    except Exception as exc:
        app.logger.error(f"[AUTH] Erro na autenticação: {exc}", exc_info=True)
        return None


def update_user_password(user_id: int, new_password: str, new_email: str = None) -> bool:
    try:
        conn = get_db()
        cursor = conn.cursor()
        password_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())
        
        # Se novo email fornecido e termina com @portoex.com.br, atualiza
        if new_email and new_email.lower().endswith("@portoex.com.br"):
            cursor.execute(
                """
                UPDATE users_new
                SET password = %s, email = %s, first_login = FALSE, 
                    updated_at = CURRENT_TIMESTAMP, last_login = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (psycopg2.Binary(password_hash), new_email.lower(), user_id),
            )
        else:
            cursor.execute(
                """
                UPDATE users_new
                SET password = %s, first_login = FALSE, updated_at = CURRENT_TIMESTAMP,
                    last_login = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (psycopg2.Binary(password_hash), user_id),
            )
        conn.commit()
        conn.close()
        return True
    except Exception as exc:
        print(f"Erro ao atualizar senha: {exc}")
        return False
            
            
def get_user_by_id(user_id: int):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT id, username, nome_completo, cargo_original,
                   departamento, role, email, is_active, nome_usuario,
                   avatar_url
            FROM users_new
            WHERE id = %s
            """,
            (user_id,),
        )
        user = cursor.fetchone()
        conn.close()
        if user:
            return dict(user)
        return None
    except Exception as exc:
        print(f"Erro ao buscar usuário: {exc}")
        return None


# --------------------------------------------------------------------------- #
# Rotas principais
# --------------------------------------------------------------------------- #
@app.route("/")
def index():
    if "user_id" in session:
        return (
            redirect(url_for("admin_dashboard"))
            if is_admin_session()
            else redirect(url_for("team_dashboard"))
        )
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if "new_password" in request.form:
            user_id = session.get("temp_user_id")
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            if not user_id:
                flash("Sessão expirada. Faça login novamente.", "error")
                return redirect(url_for("login"))

            if len(new_password) < 6:
                flash("A nova senha deve ter pelo menos 6 caracteres.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

            if new_password != confirm_password:
                flash("As senhas não coincidem.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

            # Permite atualizar email no primeiro acesso se fornecido
            new_email = request.form.get("new_email", "").strip()
            if new_email and not new_email.lower().endswith("@portoex.com.br"):
                flash("O email deve terminar com @portoex.com.br", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)
            
            if update_user_password(user_id, new_password, new_email if new_email else None):
                user = get_user_by_id(user_id)
                if user:
                    session.update(
                        {
                            "user_id": user["id"],
                            "username": user["username"],
                            "role": user["role"],
                            "email": user.get("email", ""),
                            "nome_completo": user["nome_completo"],
                            "departamento": user["departamento"],
                        }
                    )
                    session.pop("temp_user_id", None)
                    flash(
                        f"Senha atualizada! Bem-vindo, {user['nome_completo']}!",
                        "success",
                    )
                    return redirect(url_for("index"))
                flash("Não foi possível carregar o usuário.", "error")
            else:
                flash("Erro ao atualizar senha. Tente novamente.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

        identifier = (
            request.form.get("username") or request.form.get("email", "")
        ).strip()
        password = request.form.get("password", "").strip()

        if identifier and password:
            # Primeiro tenta autenticar
            user = authenticate_user(identifier, password)
            
            if user:
                # Se é primeiro acesso, permite qualquer email
                # Se não é primeiro acesso, valida email @portoex.com.br
                if not user["first_login"] and "@" in identifier:
                    # Se não é primeiro acesso e o email usado não termina com @portoex.com.br
                    if not identifier.lower().endswith("@portoex.com.br"):
                        # Verifica se o email do usuário no banco termina com @portoex.com.br
                        user_email = user.get("email", "").lower()
                        if user_email and not user_email.endswith("@portoex.com.br"):
                            flash(
                                "Use um email @portoex.com.br para acessar o sistema. "
                                "Se este é seu primeiro acesso, use seu email pessoal para definir a senha.",
                                "error",
                            )
                            return render_template(get_template("login.html"))
                
                # Primeiro acesso: redireciona para definir senha
                if user["first_login"]:
                    session["temp_user_id"] = user["id"]
                    flash(
                        f"Bem-vindo, {user['nome_completo']}! Defina uma nova senha.",
                        "info",
                    )
                    return render_template(get_template("first_login.html"), user=user)

                session.update(
                    {
                        "user_id": user["id"],
                        "username": user["username"],
                        "role": user["role"],
                        "email": user.get("email", ""),
                        "nome_completo": user["nome_completo"],
                        "departamento": user["departamento"],
                    }
                )
                flash(f"Bem-vindo de volta, {user['nome_completo']}!", "success")
                return redirect(url_for("index"))
            else:
                flash("Usuário ou senha incorretos!", "error")
        else:
            flash("Por favor, informe usuário/email e senha.", "error")

    return render_template(get_template("login.html"))


@app.route("/logout")
def logout():
    session.clear()
    flash("Logout realizado com sucesso!", "success")
    return redirect(url_for("login"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = get_user_by_id(session["user_id"])
    if not user:
        flash("Não foi possível carregar seu perfil.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        user_id = session["user_id"]
        conn = get_db()
        cursor = conn.cursor()

        errors: List[str] = []
        updates: List[str] = []
        params: List = []
        avatar_payload = None
        avatar_meta = None

        # Campos principais
        nome_completo = request.form.get("nome_completo", "").strip()
        if not nome_completo:
            errors.append("Informe seu nome completo.")
        elif nome_completo != user["nome_completo"]:
            updates.append("nome_completo = %s")
            params.append(nome_completo)

        new_username = request.form.get("username", "").strip()
        if not new_username:
            errors.append("Informe um nome de usuário.")
        elif new_username.lower() != user["username"].lower():
            cursor.execute(
                "SELECT id FROM users_new WHERE LOWER(username) = LOWER(%s) AND id <> %s",
                (new_username, user_id),
            )
            if cursor.fetchone():
                errors.append("Este nome de usuário já está em uso.")
            else:
                updates.append("username = %s")
                params.append(new_username)

        new_nome_usuario = _sanitize_optional(request.form.get("nome_usuario"))
        current_nome_usuario = _sanitize_optional(user.get("nome_usuario"))
        if new_nome_usuario != current_nome_usuario:
            if new_nome_usuario:
                cursor.execute(
                    """
                    SELECT id FROM users_new
                    WHERE LOWER(nome_usuario) = LOWER(%s) AND id <> %s
                    """,
                    (new_nome_usuario, user_id),
                )
                if cursor.fetchone():
                    errors.append("Este usuário público já está em uso.")
                else:
                    updates.append("nome_usuario = %s")
                    params.append(new_nome_usuario)
            else:
                updates.append("nome_usuario = %s")
                params.append(None)

        new_email = _sanitize_optional(request.form.get("email"))
        current_email = _sanitize_optional(user.get("email"))
        if new_email != current_email:
            if new_email:
                if not re.match(r"[^@]+@[^@]+\.[^@]+", new_email):
                    errors.append("Informe um email válido.")
                else:
                    cursor.execute(
                        "SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s) AND id <> %s",
                        (new_email, user_id),
                    )
                    if cursor.fetchone():
                        errors.append("Este email já está em uso.")
                    else:
                        updates.append("email = %s")
                        params.append(new_email.lower())
            else:
                updates.append("email = %s")
                params.append(None)

        new_departamento = _sanitize_optional(request.form.get("departamento"))
        if new_departamento != _sanitize_optional(user.get("departamento")):
            updates.append("departamento = %s")
            params.append(new_departamento)

        new_cargo = _sanitize_optional(request.form.get("cargo_original"))
        if new_cargo != _sanitize_optional(user.get("cargo_original")):
            updates.append("cargo_original = %s")
            params.append(new_cargo)

        # Upload de avatar
        avatar_file = request.files.get("avatar")
        if avatar_file and avatar_file.filename:
            if not is_allowed_avatar_file(avatar_file.filename):
                errors.append("Formato de imagem não suportado. Use PNG, JPG, JPEG, WEBP ou GIF.")
            else:
                avatar_bytes = avatar_file.read()
                if not avatar_bytes:
                    errors.append("Não foi possível ler o arquivo da foto.")
                elif len(avatar_bytes) > MAX_AVATAR_SIZE_BYTES:
                    errors.append("A foto deve ter no máximo 5MB.")
                else:
                    ext = avatar_file.filename.rsplit(".", 1)[-1].lower()
                    base_name = secure_filename(os.path.splitext(avatar_file.filename)[0]) or "avatar"
                    unique_name = f"user_{user_id}_{int(time.time())}_{base_name}.{ext}"
                    avatar_meta = (
                        avatar_bytes,
                        unique_name,
                        avatar_file.mimetype
                        or mimetypes.guess_type(avatar_file.filename)[0]
                        or "application/octet-stream",
                    )

        # Alteração de senha
        new_password = request.form.get("new_password", "").strip()
        if new_password:
            current_password = request.form.get("current_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()
            if not current_password:
                errors.append("Informe sua senha atual para definir uma nova.")
            elif len(new_password) < 6:
                errors.append("A nova senha deve ter pelo menos 6 caracteres.")
            elif new_password != confirm_password:
                errors.append("A confirmação da nova senha não confere.")
            else:
                cursor.execute("SELECT password FROM users_new WHERE id = %s", (user_id,))
                stored = cursor.fetchone()
                if not stored or not bcrypt.checkpw(
                    current_password.encode("utf-8"), _as_bytes(stored["password"])
                ):
                    errors.append("Senha atual incorreta.")
                else:
                    password_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())
                    updates.append("password = %s")
                    params.append(psycopg2.Binary(password_hash))
                    updates.append("first_login = FALSE")

        new_avatar_url = None
        if not errors and avatar_meta:
            try:
                avatar_buffer = BytesIO(avatar_meta[0])
                new_avatar_url, _ = upload_to_supabase(
                    avatar_buffer, avatar_meta[1], avatar_meta[2], folder="avatars"
                )
                updates.append("avatar_url = %s")
                params.append(new_avatar_url)
            except Exception as avatar_exc:
                app.logger.error(f"Erro ao enviar avatar: {avatar_exc}")
                errors.append("Não foi possível salvar a nova foto de perfil. Tente novamente.")

        if errors:
            for error in errors:
                flash(error, "error")
        else:
            if updates:
                updates.append("updated_at = NOW()")
                query = f"UPDATE users_new SET {', '.join(updates)} WHERE id = %s"
                params.append(user_id)
                try:
                    cursor.execute(query, params)
                    conn.commit()
                    user = refresh_session_user_cache() or get_user_by_id(user_id)
                    flash("Perfil atualizado com sucesso!", "success")
                except Exception as exc:
                    conn.rollback()
                    app.logger.error(f"Erro ao atualizar perfil do usuário {user_id}: {exc}")
                    flash("Erro ao atualizar perfil. Tente novamente.", "error")
            else:
                flash("Nenhuma alteração detectada.", "info")

        cursor.close()

    return render_template(get_template("profile.html"), user=user)


@app.route("/profile/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not current_password or not new_password or not confirm_password:
            flash("Preencha todos os campos.", "error")
            return render_template("change_password.html")

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT password FROM users_new WHERE id = %s", (session["user_id"],)
        )
        result = cursor.fetchone()
        conn.close()

        if not result or not bcrypt.checkpw(
            current_password.encode("utf-8"), _as_bytes(result["password"])
        ):
            flash("Senha atual incorreta.", "error")
            return render_template("change_password.html")

        if len(new_password) < 6:
            flash("A nova senha deve conter pelo menos 6 caracteres.", "error")
            return render_template("change_password.html")

        if new_password != confirm_password:
            flash("As senhas não coincidem.", "error")
            return render_template("change_password.html")

        if update_user_password(session["user_id"], new_password):
            flash("Senha alterada com sucesso!", "success")
            return redirect(url_for("profile"))

        flash("Erro ao alterar senha. Tente novamente.", "error")

    return render_template("change_password.html")


@app.route("/admin/dashboard")
@login_required
@admin_required
def admin_dashboard():
    users = fetch_users()
    dashboards = fetch_dashboards()
    dashboard_map = get_user_dashboard_map()
    for user in users:
        dashboard_map.setdefault(user["id"], {"ids": [], "items": []})

    selected_user_id = request.args.get("user_id", type=int)
    if selected_user_id is None and users:
        selected_user_id = users[0]["id"]

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as count FROM users_new WHERE is_active = true")
    total_users = cursor.fetchone()["count"]
    cursor.execute(
        "SELECT COUNT(*) as count FROM users_new WHERE is_active = true AND role = 'admin'"
    )
    total_admins = cursor.fetchone()["count"]
    cursor.execute("SELECT COUNT(*) as count FROM dashboards WHERE is_active = true")
    active_dashboards = cursor.fetchone()["count"]
    cursor.execute("SELECT COUNT(*) as count FROM user_dashboards")
    total_assignments = cursor.fetchone()["count"]
    cursor.execute(
        "SELECT status, created_at FROM planner_sync_logs ORDER BY created_at DESC LIMIT 1"
    )
    last_sync = cursor.fetchone()
    conn.close()

    stats = {
        "total_users": total_users,
        "total_admins": total_admins,
        "active_dashboards": active_dashboards,
        "total_assignments": total_assignments,
        "last_sync": last_sync["created_at"] if last_sync else None,
        "last_sync_status": last_sync["status"] if last_sync else None,
    }

    return render_template(
        get_template("admin_dashboard.html"),
        stats=stats,
        users=users,
        dashboards=dashboards,
        selected_user_id=selected_user_id,
        user_dashboards=dashboard_map,
        planner_enabled=planner_client.is_configured,
        planner_logs=get_recent_planner_logs(),
    )


@app.route("/admin/dashboard/permissions", methods=["POST"])
@login_required
@admin_required
def update_dashboard_permissions():
    user_id = request.form.get("user_id", type=int)
    if not user_id:
        flash("Selecione um usuário.", "error")
        return redirect(url_for("admin_dashboard"))

    dashboard_ids = request.form.getlist("dashboards")
    dashboard_ids_int = [int(d_id) for d_id in dashboard_ids]
    save_user_dashboards(user_id, dashboard_ids_int, session["user_id"])
    flash("Visibilidade atualizada com sucesso!", "success")
    return redirect(url_for("admin_dashboard", user_id=user_id))


@app.route("/admin/users")
@login_required
@admin_required
def admin_users():
    """Página de gerenciamento de usuários"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """
            SELECT id, username, nome_completo, email, nome_usuario, role, 
                   departamento, is_active, first_login, created_at
            FROM users_new
            ORDER BY nome_completo
            """
        )
        users = [dict(row) for row in cursor.fetchall()]
        
        cursor.execute(
            """
            SELECT DISTINCT departamento 
            FROM users_new 
            WHERE departamento IS NOT NULL
            ORDER BY departamento
            """
        )
        departments = [row["departamento"] for row in cursor.fetchall()]
        
        return render_template(
            get_template("admin_users.html"),
            users=users,
            departments=departments
        )
    finally:
        conn.close()


@app.route("/api/agent/knowledge/ingest-documents", methods=["POST"])
@login_required
def ingest_documents_route():
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem executar a ingestão"}), 403

    try:
        from scripts.doc_ingest import ingest_documents
        summary = ingest_documents()
        return jsonify({"success": True, "summary": summary})
    except Exception as e:
        app.logger.error(f"[AGENT-KB] Erro na ingestão de documentos: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/add", methods=["POST"])
@login_required
@admin_required
def add_user():
    """Adiciona um novo usuário"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        nome_completo = request.form.get("nome_completo", "").strip()
        email = request.form.get("email", "").strip()
        nome_usuario = request.form.get("nome_usuario", "").strip()
        role = request.form.get("role", "usuario").strip()
        password = request.form.get("password", "").strip()
        departamento = request.form.get("departamento", "").strip()
        cargo_original = request.form.get("cargo_original", "").strip()
        
        if not nome_completo or not email:
            flash("Nome completo e email são obrigatórios.", "error")
            return redirect(url_for("admin_users"))
        
        # Valida email
        if not email.lower().endswith("@portoex.com.br"):
            flash("O email deve terminar com @portoex.com.br", "error")
            return redirect(url_for("admin_users"))
        
        # Gera senha padrão se não fornecida
        if not password:
            password = "portoex123"  # Senha padrão
            first_login = True
        else:
            first_login = False
        
        # Verifica se email já existe
        cursor.execute("SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s)", (email,))
        if cursor.fetchone():
            flash("Email já cadastrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Verifica se nome_usuario já existe (se fornecido)
        if nome_usuario:
            cursor.execute("SELECT id FROM users_new WHERE LOWER(nome_usuario) = LOWER(%s)", (nome_usuario,))
            if cursor.fetchone():
                flash("Nome de usuário já cadastrado.", "error")
                return redirect(url_for("admin_users"))
        
        # Hash da senha
        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        username = email.lower()
        
        # Insere usuário
        cursor.execute(
            """
            INSERT INTO users_new (
                username, password, nome_completo, email, nome_usuario,
                role, departamento, cargo_original, is_active, first_login
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                username,
                psycopg2.Binary(password_hash),
                nome_completo,
                email.lower(),
                nome_usuario.lower() if nome_usuario else None,
                role if role in ["admin", "usuario"] else "usuario",
                departamento or None,
                cargo_original or None,
                True,
                first_login,
            ),
        )
        conn.commit()
        flash("Usuário adicionado com sucesso!", "success")
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao adicionar usuário: {exc}", exc_info=True)
        flash(f"Erro ao adicionar usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):
    """Exclui um usuário (soft delete - desativa)"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verifica se o usuário existe
        cursor.execute("SELECT id, nome_completo FROM users_new WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Não permite excluir a si mesmo
        if user_id == session.get("user_id"):
            flash("Você não pode excluir seu próprio usuário.", "error")
            return redirect(url_for("admin_users"))
        
        # Soft delete - desativa o usuário
        cursor.execute(
            "UPDATE users_new SET is_active = FALSE, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
            (user_id,),
        )
        conn.commit()
        flash(f"Usuário {user['nome_completo']} foi desativado com sucesso!", "success")
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao excluir usuário: {exc}", exc_info=True)
        flash(f"Erro ao excluir usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/users/<int:user_id>/update", methods=["POST"])
@login_required
@admin_required
def update_user(user_id):
    """Atualiza dados do usuário"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verifica se o usuário existe
        cursor.execute("SELECT id FROM users_new WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Coleta dados do formulário
        new_email = request.form.get("email", "").strip()
        new_nome_usuario = request.form.get("nome_usuario", "").strip()
        new_role = request.form.get("role", "usuario").strip()
        new_password = request.form.get("password", "").strip()
        reset_first_login = request.form.get("reset_first_login") == "on"
        
        # Novos campos
        new_nome_completo = request.form.get("nome_completo", "").strip()
        new_cargo = request.form.get("cargo_original", "").strip()
        new_departamento = request.form.get("departamento", "").strip()
        new_is_active = request.form.get("is_active") == "true"
        
        # Valida email
        if new_email and not new_email.lower().endswith("@portoex.com.br"):
            flash("O email deve terminar com @portoex.com.br", "error")
            return redirect(url_for("admin_users"))
        
        # Atualiza campos
        updates = []
        params = []
        
        if new_nome_completo:
            updates.append("nome_completo = %s")
            params.append(new_nome_completo)
            
        # Cargo e departamento podem ser vazios
        updates.append("cargo_original = %s")
        params.append(new_cargo)
        
        updates.append("departamento = %s")
        params.append(new_departamento)
        
        # Status
        updates.append("is_active = %s")
        params.append(new_is_active)
        
        if new_email:
            updates.append("email = %s")
            params.append(new_email.lower())
        
        if new_nome_usuario:
            updates.append("nome_usuario = %s")
            params.append(new_nome_usuario.lower())
        
        if new_role in ["admin", "usuario"]:
            updates.append("role = %s")
            params.append(new_role)
        
        if new_password:
            password_hash = bcrypt.hashpw(
                new_password.encode("utf-8"), bcrypt.gensalt()
            )
            updates.append("password = %s")
            params.append(psycopg2.Binary(password_hash))
            updates.append("first_login = FALSE")
        
        if reset_first_login:
            updates.append("first_login = TRUE")
        
        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(user_id)
            
            query = f"UPDATE users_new SET {', '.join(updates)} WHERE id = %s"
            cursor.execute(query, params)
            conn.commit()
            flash("Usuário atualizado com sucesso!", "success")
        else:
            flash("Nenhuma alteração foi feita.", "info")
        
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao atualizar usuário: {exc}", exc_info=True)
        flash(f"Erro ao atualizar usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/planner/sync", methods=["POST"])
@login_required
@admin_required
def admin_planner_sync():
    try:
        success_count, errors = sync_dashboards_to_planner()
        if success_count:
            flash(
                f"Agenda enviada para {success_count} usuário(s) no Planner.",
                "success",
            )
        if errors:
            flash("Falha para: " + "; ".join(errors), "error")
    except PlannerIntegrationError as exc:
        flash(str(exc), "error")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/environments")
@login_required
@admin_required
def admin_environments():
    """Página para gerenciar ambientes do CD"""
    return render_template(get_template("admin_environments.html"))


@app.route("/admin/dashboards/add", methods=["GET", "POST"])
@login_required
@admin_required
def admin_add_dashboard():
    """Adiciona ou edita um dashboard do BI"""
    if request.method == "POST":
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            dashboard_id = request.form.get("dashboard_id", type=int)
            title = request.form.get("title", "").strip()
            description = request.form.get("description", "").strip()
            category = request.form.get("category", "").strip()
            embed_code = request.form.get("embed_code", "").strip()
            display_order = request.form.get("display_order", type=int) or 0
            
            if not title or not embed_code:
                flash("Nome e código de incorporação são obrigatórios.", "error")
                return redirect(url_for("admin_add_dashboard", dashboard_id=dashboard_id) if dashboard_id else url_for("admin_add_dashboard"))
            
            # Extrai URL do iframe
            url_match = re.search(r'src="([^"]+)"', embed_code)
            if not url_match:
                flash("Código de incorporação inválido. Deve conter um iframe com src.", "error")
                return redirect(url_for("admin_add_dashboard", dashboard_id=dashboard_id) if dashboard_id else url_for("admin_add_dashboard"))
            
            embed_url = url_match.group(1)
            
            # Gera slug do título
            slug = title.lower().replace(" ", "-").replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
            slug = re.sub(r'[^a-z0-9-]', '', slug)
            
            if dashboard_id:
                # Atualiza dashboard existente
                cursor.execute(
                    """
                    UPDATE dashboards
                    SET title = %s, description = %s, category = %s, 
                        embed_url = %s, display_order = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (title, description, category, embed_url, display_order, dashboard_id),
                )
                flash("Dashboard atualizado com sucesso!", "success")
            else:
                # Cria novo dashboard
                cursor.execute(
                    """
                    INSERT INTO dashboards (slug, title, description, category, embed_url, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (slug) DO UPDATE SET
                        title = EXCLUDED.title,
                        description = EXCLUDED.description,
                        category = EXCLUDED.category,
                        embed_url = EXCLUDED.embed_url,
                        display_order = EXCLUDED.display_order,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (slug, title, description, category, embed_url, display_order),
                )
                flash("Dashboard adicionado com sucesso!", "success")
            
            conn.commit()
            return redirect(url_for("admin_dashboard"))
        except Exception as exc:
            conn.rollback()
            app.logger.error(f"Erro ao salvar dashboard: {exc}", exc_info=True)
            flash(f"Erro ao salvar dashboard: {exc}", "error")
            return redirect(url_for("admin_add_dashboard"))
        finally:
            conn.close()
    
    # GET: mostra formulário
    dashboard_id = request.args.get("dashboard_id", type=int)
    dashboard = None
    
    if dashboard_id:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM dashboards WHERE id = %s", (dashboard_id,))
        dashboard = cursor.fetchone()
        conn.close()
        
        if not dashboard:
            flash("Dashboard não encontrado.", "error")
            return redirect(url_for("admin_dashboard"))
    
    return render_template(get_template("admin_add_dashboard.html"), dashboard=dashboard)


@app.route("/favicon.ico")
def favicon():
    """Serve o favicon"""
    return send_from_directory(
        app.static_folder,
        "test_favicon.ico",
        mimetype="image/x-icon"
    )


@app.route("/team/dashboard")
@app.route("/dashboards")
@login_required
def team_dashboard():
    show_all = is_admin_session() and request.args.get("all") == "1"
    conn = None

    if show_all and is_admin_session():
        dashboards = fetch_dashboards()
    else:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT d.title, d.description, d.category, d.embed_url
            FROM dashboards d
            JOIN user_dashboards ud ON ud.dashboard_id = d.id
            WHERE ud.user_id = %s AND d.is_active = true
            ORDER BY d.display_order, d.title
            """,
            (session["user_id"],),
        )
        dashboards = cursor.fetchall()
        conn.close()

    today_label = datetime.now().strftime("%d/%m/%Y")
    return render_template(
        get_template("team_dashboard.html"),
        user=session.get("nome_completo", "Usuário"),
        dashboards=dashboards,
        today=today_label,
        show_all=show_all,
        is_admin=is_admin_session(),
    )


@app.route("/cd/facilities")
@login_required
def cd_facilities():
    """Página para visualizar planta 3D do CD"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        # Buscar ambientes ordenados por display_order
        cursor.execute("""
            SELECT id, code, name, description, icon, floor
            FROM environments 
            WHERE is_active = true 
            ORDER BY display_order ASC
        """)
        environments = [dict(row) for row in cursor.fetchall()]
        
        # Buscar recursos (modelos 3D, fotos, plantas) para cada ambiente
        # Isso é importante para que o JS saiba o que carregar
        cursor.execute("""
            SELECT environment_id, resource_type, file_url, is_primary
            FROM environment_resources
            ORDER BY is_primary DESC
        """)
        resources = [dict(row) for row in cursor.fetchall()]
        
        # Agrupar recursos por ambiente
        env_resources = {}
        for res in resources:
            env_id = res['environment_id']
            if env_id not in env_resources:
                env_resources[env_id] = []
            env_resources[env_id].append(res)
            
        # Adicionar recursos aos ambientes
        for env in environments:
            env['resources'] = env_resources.get(env['id'], [])
            
    except Exception as e:
        app.logger.error(f"Erro ao carregar ambientes: {str(e)}")
        environments = []
    finally:
        conn.close()
        
    return render_template(get_template("cd_facilities.html"), environments=environments)


@app.route("/cd/booking")
@login_required
def cd_booking():
    """Página para agendar salas de reunião"""
    return render_template(get_template("cd_booking.html"))


@app.route("/api/room-bookings", methods=["GET", "POST"])
@login_required
def room_bookings_api():
    """API para listar e criar agendamentos de salas"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute("""
                SELECT 
                    rb.id, rb.room, rb.title, rb.date, rb.start_time, rb.end_time,
                    rb.participants, rb.subject, rb.created_at,
                    u.nome_completo as user_name
                FROM room_bookings rb
                JOIN users_new u ON rb.user_id = u.id
                WHERE rb.is_active = true
                ORDER BY rb.date DESC, rb.start_time DESC
            """)
            bookings = cursor.fetchall()
            
            # Serializar objetos de data/hora
            result = []
            for row in bookings:
                booking = dict(row)
                booking['date'] = str(booking['date'])
                booking['start_time'] = str(booking['start_time'])
                booking['end_time'] = str(booking['end_time'])
                booking['created_at'] = str(booking['created_at'])
                result.append(booking)
                
            return jsonify(result)
        
        elif request.method == "POST":
            data = request.get_json()
            
            required_fields = ['room', 'title', 'date', 'start_time', 'end_time', 'participants', 'subject']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            user_id = session.get('user_id')
            if not user_id:
                app.logger.error("Tentativa de agendamento sem user_id na sessão")
                return jsonify({"error": "Sessão inválida. Faça login novamente."}), 401
                
            app.logger.info(f"Tentando criar agendamento: User={user_id}, Room={data['room']}, Date={data['date']}")
            
            cursor.execute("""
                SELECT id FROM room_bookings
                WHERE room = %s AND date = %s AND is_active = true
                AND (
                    (start_time <= %s AND end_time > %s) OR
                    (start_time < %s AND end_time >= %s) OR
                    (start_time >= %s AND end_time <= %s)
                )
            """, (
                data['room'], data['date'],
                data['start_time'], data['start_time'],
                data['end_time'], data['end_time'],
                data['start_time'], data['end_time']
            ))
            
            conflict = cursor.fetchone()
            if conflict:
                return jsonify({"error": "Já existe um agendamento neste horário para esta sala"}), 409
            
            cursor.execute("""
                INSERT INTO room_bookings 
                (user_id, room, title, date, start_time, end_time, participants, subject)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                user_id,
                data['room'],
                data['title'],
                data['date'],
                data['start_time'],
                data['end_time'],
                data['participants'],
                data['subject']
            ))
            
            booking_id = cursor.fetchone()['id']
            conn.commit()
            
            app.logger.info(f"Agendamento criado com sucesso: ID={booking_id}")
            return jsonify({"success": True, "id": booking_id}), 201
    
    except Exception as e:
        conn.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Erro ao criar agendamento: {str(e)}\n{error_details}")
        # Retorna o erro detalhado apenas em debug, ou uma mensagem genérica
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500
    finally:
        conn.close()


@app.route("/api/room-bookings/<int:booking_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def room_booking_detail_api(booking_id):
    """API para obter, atualizar ou deletar um agendamento específico"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute("""
                SELECT 
                    rb.id, rb.room, rb.title, rb.date, rb.start_time, rb.end_time,
                    rb.participants, rb.subject, rb.created_at, rb.user_id,
                    u.nome_completo as user_name
                FROM room_bookings rb
                JOIN users_new u ON rb.user_id = u.id
                WHERE rb.id = %s AND rb.is_active = true
            """, (booking_id,))
            
            booking = cursor.fetchone()
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            # Serializar
            result = dict(booking)
            result['date'] = str(result['date'])
            result['start_time'] = str(result['start_time'])
            result['end_time'] = str(result['end_time'])
            result['created_at'] = str(result['created_at'])
            
            return jsonify(result)
        
        elif request.method == "PUT":
            cursor.execute(
                "SELECT user_id FROM room_bookings WHERE id = %s AND is_active = true",
                (booking_id,)
            )
            booking = cursor.fetchone()
            
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            if booking['user_id'] != session['user_id'] and session.get('role') != 'admin':
                return jsonify({"error": "Sem permissão para editar este agendamento"}), 403
            
            data = request.get_json()
            
            cursor.execute("""
                UPDATE room_bookings
                SET room = %s, title = %s, date = %s, start_time = %s, 
                    end_time = %s, participants = %s, subject = %s
                WHERE id = %s
            """, (
                data.get('room'),
                data.get('title'),
                data.get('date'),
                data.get('start_time'),
                data.get('end_time'),
                data.get('participants'),
                data.get('subject'),
                booking_id
            ))
            
            conn.commit()
            return jsonify({"success": True})
        
        elif request.method == "DELETE":
            cursor.execute(
                "SELECT user_id FROM room_bookings WHERE id = %s AND is_active = true",
                (booking_id,)
            )
            booking = cursor.fetchone()
            
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            if booking['user_id'] != session['user_id'] and session.get('role') != 'admin':
                return jsonify({"error": "Sem permissão para excluir este agendamento"}), 403
            
            cursor.execute(
                "UPDATE room_bookings SET is_active = false WHERE id = %s",
                (booking_id,)
            )
            
            conn.commit()
            return jsonify({"success": True})
    
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/3d-model/<model_type>", methods=['GET', 'OPTIONS'])
def proxy_3d_model(model_type):
    """Proxy otimizado para modelos 3D com streaming e CORS"""
    
    # Responder a preflight OPTIONS
    if request.method == 'OPTIONS':
        response = app.make_default_options_response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    
    app.logger.info(f"[3D PROXY] Requisitando modelo: {model_type}")
    
    urls = {
        'glb': 'https://github.com/anaissiabraao/GeRot/releases/download/v1.0-3d-models/Cd_front_12_50_53.glb',
        'fbx': 'https://github.com/anaissiabraao/GeRot/releases/download/v1.0-3d-models/Cd_front_12_10_17.fbx'
    }
    
    if model_type not in urls:
        app.logger.error(f"[3D PROXY] Modelo inválido: {model_type}")
        return jsonify({"error": "Modelo inválido"}), 404
    
    try:
        # Streaming otimizado: chunks grandes + timeout longo
        app.logger.info(f"[3D PROXY] Baixando: {urls[model_type]}")
        
        # Fazer requisição (sem context manager para não fechar antes do generator terminar)
        # Timeout de 300s (5min) para arquivos grandes
        r = requests.get(urls[model_type], stream=True, timeout=(30, 300))
        r.raise_for_status()
        
        content_length = r.headers.get('content-length', '0')
        content_type = r.headers.get('content-type', 'application/octet-stream')
        
        app.logger.info(f"[3D PROXY] Tamanho do arquivo: {content_length} bytes ({int(content_length)/(1024*1024):.2f} MB)")
        
        def generate():
            """Generator para streaming eficiente com keep-alive"""
            try:
                bytes_sent = 0
                # Chunks de 256KB para velocidade máxima
                for chunk in r.iter_content(chunk_size=262144):
                    if chunk:
                        bytes_sent += len(chunk)
                        yield chunk
                        # Log a cada 10MB para monitorar progresso
                        if bytes_sent % (10 * 1024 * 1024) < 262144:
                            app.logger.info(f"[3D PROXY] {bytes_sent/(1024*1024):.1f}MB enviados")
            finally:
                # Fechar conexão após streaming completo
                r.close()
                app.logger.info(f"[3D PROXY] Streaming concluído: {bytes_sent/(1024*1024):.1f}MB para {model_type}")
        
        # Retornar resposta com CORS, Content-Length e streaming
        from flask import Response
        response = Response(generate(), mimetype=content_type, direct_passthrough=True)
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        response.headers['Content-Length'] = content_length
        response.headers['Cache-Control'] = 'public, max-age=86400'
        response.headers['Connection'] = 'keep-alive'
        response.headers['Keep-Alive'] = 'timeout=300'
        
        app.logger.info(f"[3D PROXY] Streaming iniciado para {model_type}")
        return response
        
    except requests.exceptions.Timeout:
        app.logger.error(f"[3D PROXY] Timeout ao baixar {model_type}")
        return jsonify({"error": "Timeout ao baixar modelo"}), 504
    except requests.exceptions.RequestException as e:
        app.logger.error(f"[3D PROXY] Erro de rede: {str(e)}")
        return jsonify({"error": f"Erro ao baixar modelo: {str(e)}"}), 502
    except Exception as e:
        app.logger.error(f"[3D PROXY] Erro inesperado: {str(e)}")
        return jsonify({"error": str(e)}), 500


# --------------------------------------------------------------------------- #
# API pública básica
# --------------------------------------------------------------------------- #
class UsersAPI(Resource):
    def get(self, user_id=None):
        # Validação de admin para acessar lista de usuários
        if session.get("role") != "admin":
            return jsonify({"error": "Acesso restrito aos administradores"}), 403
        
        conn = get_db()
        cursor = conn.cursor()
        try:
            if user_id:
                cursor.execute(
                    "SELECT id, username, nome_completo, role, departamento FROM users_new WHERE id = %s AND is_active = true",
                    (user_id,),
                )
                user = cursor.fetchone()
                if not user:
                    return jsonify({"error": "User not found"}), 404
                return jsonify({"user": dict(user)})

            cursor.execute(
                "SELECT id, username, nome_completo, role, departamento FROM users_new WHERE is_active = true"
            )
            users = [dict(row) for row in cursor.fetchall()]
            return jsonify({"users": users})
        finally:
            conn.close()


api.add_resource(UsersAPI, "/api/users", "/api/users/<int:user_id>")


# --------------------------------------------------------------------------- #
# API para gerenciar ambientes do CD
# --------------------------------------------------------------------------- #
@app.route("/api/environments", methods=["GET", "POST"])
@login_required
def environments_api():
    """API para listar e criar ambientes"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute(""" 
                SELECT 
                    e.id, e.code, e.name, e.description, e.icon, 
                    e.capacity, e.area_m2, e.floor, e.is_active,
                    e.display_order, e.created_at,
                    COUNT(DISTINCT er.id) as resource_count,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'model_3d' THEN er.id END) as models_3d,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'plant_2d' THEN er.id END) as plants_2d,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'photo' THEN er.id END) as photos
                FROM environments e
                LEFT JOIN environment_resources er ON e.id = er.environment_id
                WHERE e.is_active = true
                GROUP BY e.id, e.code, e.name, e.description, e.icon, 
                         e.capacity, e.area_m2, e.floor, e.is_active,
                         e.display_order, e.created_at
                ORDER BY e.display_order, e.name
            """)
            environments = cursor.fetchall()
            return jsonify([dict(row) for row in environments])
        
        elif request.method == "POST":
            # Apenas admins podem criar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem criar ambientes"}), 403
            
            data = request.get_json()
            
            required_fields = ['code', 'name']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            cursor.execute(""" 
                INSERT INTO environments 
                (code, name, description, icon, capacity, area_m2, floor, display_order)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                data['code'],
                data['name'],
                data.get('description', ''),
                data.get('icon', 'fas fa-building'),
                data.get('capacity'),
                data.get('area_m2'),
                data.get('floor', 1),
                data.get('display_order', 0)
            ))
            
            environment_id = cursor.fetchone()['id']
            conn.commit()
            
            return jsonify({"success": True, "id": environment_id}), 201
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar ambientes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/environments/<int:environment_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def environment_detail_api(environment_id):
    """API para obter, atualizar ou deletar um ambiente específico"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute(""" 
                SELECT 
                    e.*, 
                    s.camera_position_x, s.camera_position_y, s.camera_position_z,
                    s.camera_target_x, s.camera_target_y, s.camera_target_z,
                    s.model_scale, s.rotation_speed, s.enable_shadows,
                    s.background_color, s.grid_size
                FROM environments e
                LEFT JOIN environment_3d_settings s ON e.id = s.environment_id
                WHERE e.id = %s AND e.is_active = true
            """, (environment_id,))
            
            environment = cursor.fetchone()
            if not environment:
                return jsonify({"error": "Ambiente não encontrado"}), 404
            
            # Buscar recursos associados
            cursor.execute(""" 
                SELECT * FROM environment_resources 
                WHERE environment_id = %s 
                ORDER BY resource_type, display_order
            """, (environment_id,))
            resources = cursor.fetchall()
            
            result = dict(environment)
            result['resources'] = [dict(r) for r in resources]
            
            return jsonify(result)
        
        elif request.method == "PUT":
            # Apenas admins podem atualizar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem editar ambientes"}), 403
            
            data = request.get_json()
            
            # Atualizar ambiente
            cursor.execute(""" 
                UPDATE environments 
                SET name = %s, description = %s, icon = %s, 
                    capacity = %s, area_m2 = %s, floor = %s, 
                    display_order = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND is_active = true
            """, (
                data.get('name'),
                data.get('description'),
                data.get('icon'),
                data.get('capacity'),
                data.get('area_m2'),
                data.get('floor'),
                data.get('display_order'),
                environment_id
            ))
            
            # Atualizar configurações 3D se fornecidas
            if '3d_settings' in data:
                settings = data['3d_settings']
                
                # Verificar se já existe configuração
                cursor.execute(
                    "SELECT id FROM environment_3d_settings WHERE environment_id = %s",
                    (environment_id,)
                )
                exists = cursor.fetchone()
                
                if exists:
                    cursor.execute(""" 
                        UPDATE environment_3d_settings
                        SET camera_position_x = %s, camera_position_y = %s, camera_position_z = %s,
                            camera_target_x = %s, camera_target_y = %s, camera_target_z = %s,
                            model_scale = %s, rotation_speed = %s, enable_shadows = %s,
                            background_color = %s, grid_size = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE environment_id = %s
                    """, (
                        settings.get('camera_position_x', 5),
                        settings.get('camera_position_y', 5),
                        settings.get('camera_position_z', 5),
                        settings.get('camera_target_x', 0),
                        settings.get('camera_target_y', 0),
                        settings.get('camera_target_z', 0),
                        settings.get('model_scale', 1.0),
                        settings.get('rotation_speed', 0.01),
                        settings.get('enable_shadows', True),
                        settings.get('background_color', '#1a1a2e'),
                        settings.get('grid_size', 20),
                        environment_id
                    ))
                else:
                    cursor.execute(""" 
                        INSERT INTO environment_3d_settings
                        (environment_id, camera_position_x, camera_position_y, camera_position_z,
                         camera_target_x, camera_target_y, camera_target_z,
                         model_scale, rotation_speed, enable_shadows, background_color, grid_size)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        environment_id,
                        settings.get('camera_position_x', 5),
                        settings.get('camera_position_y', 5),
                        settings.get('camera_position_z', 5),
                        settings.get('camera_target_x', 0),
                        settings.get('camera_target_y', 0),
                        settings.get('camera_target_z', 0),
                        settings.get('model_scale', 1.0),
                        settings.get('rotation_speed', 0.01),
                        settings.get('enable_shadows', True),
                        settings.get('background_color', '#1a1a2e'),
                        settings.get('grid_size', 20)
                    ))
            
            conn.commit()
            return jsonify({"success": True})
        
        elif request.method == "DELETE":
            # Apenas admins podem deletar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem excluir ambientes"}), 403
            
            # Soft delete - apenas marca como inativo
            cursor.execute(""" 
                UPDATE environments 
                SET is_active = false, updated_at = CURRENT_TIMESTAMP 
                WHERE id = %s
            """, (environment_id,))
            
            conn.commit()
            return jsonify({"success": True})
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar ambiente {environment_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/environments/<int:environment_id>/resources", methods=["GET", "POST"])
@login_required
def environment_resources_api(environment_id):
    """API para gerenciar recursos (imagens, modelos 3D) de um ambiente"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            resource_type = request.args.get('type')
            
            query = """
                SELECT * FROM environment_resources 
                WHERE environment_id = %s
            """
            params = [environment_id]
            
            if resource_type:
                query += " AND resource_type = %s"
                params.append(resource_type)
            
            query += " ORDER BY display_order, created_at DESC"
            
            cursor.execute(query, params)
            resources = cursor.fetchall()
            
            return jsonify([dict(r) for r in resources])
        
        elif request.method == "POST":
            # Apenas admins podem adicionar recursos
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem adicionar recursos"}), 403
            
            data = request.get_json()
            
            required_fields = ['resource_type', 'file_name', 'file_url']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            # Se marcar como primário, desmarcar outros do mesmo tipo
            if data.get('is_primary'):
                cursor.execute(""" 
                    UPDATE environment_resources 
                    SET is_primary = false 
                    WHERE environment_id = %s AND resource_type = %s
                """, (environment_id, data['resource_type']))
            
            cursor.execute(""" 
                INSERT INTO environment_resources
                (environment_id, resource_type, file_name, file_url, file_size, 
                 mime_type, description, is_primary, display_order, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                environment_id,
                data['resource_type'],
                data['file_name'],
                data['file_url'],
                data.get('file_size'),
                data.get('mime_type'),
                data.get('description'),
                data.get('is_primary', False),
                data.get('display_order', 0),
                session.get('user_id')
            ))
            
            resource_id = cursor.fetchone()['id']
            conn.commit()
            
            return jsonify({"success": True, "id": resource_id}), 201
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar recursos do ambiente {environment_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


def get_supabase_config():
    """Recupera configurações do Supabase das variáveis de ambiente"""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    return url, key


def upload_to_supabase(file_obj, filename, content_type, folder="environments"):
    """Realiza upload de arquivo para o Supabase Storage"""
    url, key = get_supabase_config()
    
    if not url or not key:
        # Fallback para desenvolvimento local ou erro
        app.logger.error("Supabase credentials not found")
        raise Exception("Serviço de armazenamento não configurado")
    
    # Limpar URL base
    url = url.rstrip('/')
    bucket = "environment-assets"
    
    folder = folder.strip("/") if folder else "environments"
    storage_path = f"{folder}/{filename}"
    
    # Endpoint da API de Storage
    # POST /storage/v1/object/{bucket}/{path}
    api_url = f"{url}/storage/v1/object/{bucket}/{storage_path}"
    
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": content_type,
        "x-upsert": "true"
    }
    
    # Ler conteúdo do arquivo
    file_content = file_obj.read()
    
    try:
        response = requests.post(api_url, data=file_content, headers=headers)
        
        if response.status_code not in [200, 201]:
            # Se falhar, tentar criar o bucket e tentar novamente?
            # Por simplicidade, assumimos que o bucket existe.
            # Se o erro for 404 (bucket not found), logar erro específico.
            error_msg = f"Supabase Upload Failed ({response.status_code}): {response.text}"
            app.logger.error(error_msg)
            raise Exception("Falha no upload para o storage remoto")
            
        # Retornar URL pública
        # {supabase_url}/storage/v1/object/public/{bucket}/{path}
        public_url = f"{url}/storage/v1/object/public/{bucket}/{storage_path}"
        return public_url, len(file_content)
        
    except requests.exceptions.RequestException as e:
        app.logger.error(f"Erro de conexão com Supabase: {e}")
        raise Exception("Erro de conexão com serviço de storage")


@app.route("/api/environments/<int:environment_id>/upload", methods=["POST"])
@login_required
def environment_upload_api(environment_id):
    """API para upload de arquivos de ambiente"""
    # Validar permissão (admin/manager)
    if session.get("role") not in ["admin", "manager"]:
        return jsonify({"error": "Permissão negada"}), 403

    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
        
    file = request.files["file"]
    
    if file.filename == "":
        return jsonify({"error": "Nome de arquivo inválido"}), 400
        
    if file:
        try:
            filename = secure_filename(file.filename)
            
            # Adicionar timestamp para evitar colisão de nomes
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            unique_filename = f"{environment_id}_{timestamp}_{filename}"
            
            # Detectar tipo de recurso baseado na extensão
            ext = os.path.splitext(filename)[1].lower()
            mime_type = file.mimetype or mimetypes.guess_type(filename)[0]
            
            resource_type = "document"
            if ext in ['.jpg', '.jpeg', '.png', '.webp']:
                resource_type = "photo"
            elif ext in ['.glb', '.gltf', '.fbx', '.obj']:
                resource_type = "model_3d"
            elif ext in ['.pdf']:
                # PDFs podem ser plantas ou docs
                resource_type = "plant_2d" # Assumindo planta por padrão para PDF neste contexto
            
            # Realizar upload
            public_url, file_size = upload_to_supabase(file, unique_filename, mime_type)
            
            # Salvar no banco
            conn = get_db()
            cursor = conn.cursor()
            
            # Verificar se já existe um recurso primário deste tipo
            cursor.execute("""
                SELECT id FROM environment_resources 
                WHERE environment_id = %s AND resource_type = %s AND is_primary = true
            """, (environment_id, resource_type))
            has_primary = cursor.fetchone() is not None
            
            # Inserir novo recurso
            cursor.execute("""
                INSERT INTO environment_resources
                (environment_id, resource_type, file_name, file_url, file_size, 
                 mime_type, is_primary, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                environment_id,
                resource_type,
                filename,
                public_url,
                file_size,
                mime_type,
                not has_primary, # Se não tem primário, este será o primeiro
                session.get('user_id')
            ))
            
            resource_id = cursor.fetchone()['id']
            conn.commit()
            conn.close()
            
            return jsonify({
                "success": True, 
                "id": resource_id, 
                "url": public_url,
                "type": resource_type
            }), 201
            
        except Exception as e:
            app.logger.error(f"Erro no upload: {e}")
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Erro desconhecido"}), 500


@app.route("/api/resources/<int:resource_id>", methods=["DELETE"])
@login_required
def delete_resource_api(resource_id):
    """API para excluir um recurso"""
    if session.get("role") not in ["admin", "manager"]:
        return jsonify({"error": "Permissão negada"}), 403
        
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se existe
        cursor.execute("SELECT id FROM environment_resources WHERE id = %s", (resource_id,))
        if not cursor.fetchone():
            return jsonify({"error": "Recurso não encontrado"}), 404
            
        # Deletar do banco
        cursor.execute("DELETE FROM environment_resources WHERE id = %s", (resource_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# Rotas do Agente IA
# --------------------------------------------------------------------------- #
@app.route("/agent")
@login_required
@admin_required
def agent_page():
    """Página principal do Agente IA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar tipos de RPA com contagem de RPAs
        cursor.execute("""
            SELECT t.id, t.name, t.description, t.icon, t.is_active,
                   COUNT(r.id) as rpa_count
            FROM agent_rpa_types t
            LEFT JOIN agent_rpas r ON r.rpa_type_id = t.id
            WHERE t.is_active = true 
            GROUP BY t.id, t.name, t.description, t.icon, t.is_active
            ORDER BY t.name
        """)
        rpa_types = [dict(row) for row in cursor.fetchall()]
        
        # Buscar fontes de dados
        cursor.execute("""
            SELECT id, name, description, source_type 
            FROM agent_data_sources 
            WHERE is_active = true 
            ORDER BY name
        """)
        data_sources = [dict(row) for row in cursor.fetchall()]
        
        # Buscar RPAs do usuário
        cursor.execute("""
            SELECT r.id, r.name, r.status, r.priority, r.created_at,
                   t.name as type_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.created_by = %s
            ORDER BY r.created_at DESC
            LIMIT 20
        """, (session['user_id'],))
        rpas = [dict(row) for row in cursor.fetchall()]
        
        # Buscar dashboards gerados pelo usuário
        cursor.execute("""
            SELECT id, title, category, status, result_url, result_data, created_at
            FROM agent_dashboard_requests
            WHERE created_by = %s
            ORDER BY created_at DESC
            LIMIT 20
        """, (session['user_id'],))
        generated_dashboards = [dict(row) for row in cursor.fetchall()]
        
        # Estatísticas de RPA
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE status = 'pending') as pending,
                COUNT(*) FILTER (WHERE status = 'running') as running,
                COUNT(*) FILTER (WHERE status = 'completed') as completed,
                COUNT(*) FILTER (WHERE status = 'failed') as failed
            FROM agent_rpas
            WHERE created_by = %s
        """, (session['user_id'],))
        stats = dict(cursor.fetchone())
        
        # Estatísticas de Dashboard
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE status = 'pending') as pending,
                COUNT(*) FILTER (WHERE status = 'processing') as processing,
                COUNT(*) FILTER (WHERE status = 'completed') as completed
            FROM agent_dashboard_requests
            WHERE created_by = %s
        """, (session['user_id'],))
        dashboard_stats = dict(cursor.fetchone())
        
        # Buscar templates de dashboard do usuário
        dashboard_templates = []
        try:
            cursor.execute("""
                SELECT id, title, description, category, is_published, thumbnail_url, created_at
                FROM agent_dashboard_templates
                WHERE created_by = %s
                ORDER BY updated_at DESC
                LIMIT 20
            """, (session['user_id'],))
            dashboard_templates = [dict(row) for row in cursor.fetchall()]
        except Exception:
            pass  # Tabela pode não existir ainda
        
        return render_template(
            get_template("agent.html"),
            rpa_types=rpa_types,
            data_sources=data_sources,
            rpas=rpas,
            generated_dashboards=generated_dashboards,
            stats=stats,
            dashboard_stats=dashboard_stats,
            dashboard_templates=dashboard_templates
        )
        
    except Exception as e:
        app.logger.error(f"[AGENT] Erro ao carregar página: {e}")
        # Se as tabelas não existem, mostrar página com dados vazios
        return render_template(
            get_template("agent.html"),
            rpa_types=[],
            data_sources=[],
            rpas=[],
            generated_dashboards=[],
            stats={'pending': 0, 'running': 0, 'completed': 0, 'failed': 0},
            dashboard_stats={'pending': 0, 'processing': 0, 'completed': 0},
            dashboard_templates=[]
        )
    finally:
        conn.close()


@app.route("/api/agent/rpa", methods=["POST"])
@login_required
def create_rpa():
    """API para criar uma nova automação RPA."""
    data = request.get_json()
    
    if not data.get('name') or not data.get('description'):
        return jsonify({"error": "Nome e descrição são obrigatórios"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Validar parâmetros JSON se fornecido
        parameters = None
        if data.get('parameters'):
            try:
                import json
                parameters = json.loads(data['parameters']) if isinstance(data['parameters'], str) else data['parameters']
            except:
                return jsonify({"error": "Parâmetros JSON inválidos"}), 400
        
        cursor.execute("""
            INSERT INTO agent_rpas (name, description, rpa_type_id, priority, frequency, parameters, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['name'],
            data['description'],
            data.get('rpa_type') or None,
            data.get('priority', 'medium'),
            data.get('frequency', 'once'),
            psycopg2.extras.Json(parameters) if parameters else None,
            session['user_id']
        ))
        
        rpa_id = cursor.fetchone()['id']
        conn.commit()
        
        # Log da ação
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('create', 'rpa', %s, %s, %s)
        """, (rpa_id, session['user_id'], psycopg2.extras.Json({'name': data['name']})))
        conn.commit()
        
        return jsonify({"success": True, "id": rpa_id}), 201
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT] Erro ao criar RPA: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>", methods=["DELETE"])
@login_required
def delete_rpa(rpa_id):
    """API para excluir uma automação RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se pertence ao usuário ou se é admin
        cursor.execute("SELECT created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_rpas WHERE id = %s", (rpa_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen", methods=["GET"])
@login_required
def list_dashboard_gen():
    """Lista solicitações de geração de dashboard com filtros opcionais.
    Suporta filtros via query string: status=completed|pending|failed e has_data=true.
    Retorna itens com id, title, status, row_count e updated_at.
    """
    status = (request.args.get('status') or '').strip().lower()
    has_data = (request.args.get('has_data') or '').strip().lower() == 'true'

    conn = get_db()
    cursor = conn.cursor()

    try:
        where_clauses = []
        params = []

        # Restringir por usuário (a menos que admin)
        if session.get('role') != 'admin':
            where_clauses.append("created_by = %s")
            params.append(session['user_id'])

        # Filtro de status opcional
        if status in ('completed', 'pending', 'failed'):
            where_clauses.append("status = %s")
            params.append(status)

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ''

        cursor.execute(f"""
            SELECT id, title, status, result_data, updated_at
            FROM agent_dashboard_requests
            {where_sql}
            ORDER BY updated_at DESC
            LIMIT 100
        """, tuple(params))

        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            rd = r.get('result_data') or {}
            data = rd.get('data') or []
            row_count = len(data) if isinstance(data, list) else 0
            if has_data and row_count == 0:
                continue
            items.append({
                'id': r['id'],
                'title': r['title'],
                'status': r['status'],
                'row_count': row_count,
                'updated_at': r.get('updated_at').isoformat() if r.get('updated_at') else None
            })

        return jsonify({'items': items}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen", methods=["POST"])
@login_required
def create_dashboard_gen():
    """API para solicitar geração de dashboard."""
    data = request.get_json()
    
    if not data.get('title') or not data.get('description'):
        return jsonify({"error": "Título e descrição são obrigatórios"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Validar filtros JSON se fornecido
        filters = None
        if data.get('filters'):
            try:
                import json
                filters = json.loads(data['filters']) if isinstance(data['filters'], str) else data['filters']
            except:
                return jsonify({"error": "Filtros JSON inválidos"}), 400
        
        chart_types = data.get('chart_types', [])
        
        cursor.execute("""
            INSERT INTO agent_dashboard_requests 
            (title, description, category, data_source_id, chart_types, filters, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['title'],
            data['description'],
            data.get('category', 'Outros'),
            data.get('data_source') or None,
            chart_types,
            psycopg2.extras.Json(filters) if filters else None,
            session['user_id']
        ))
        
        request_id = cursor.fetchone()['id']
        conn.commit()
        
        # Log da ação
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('create', 'dashboard_request', %s, %s, %s)
        """, (request_id, session['user_id'], psycopg2.extras.Json({'title': data['title']})))
        conn.commit()
        
        return jsonify({"success": True, "id": request_id}), 201
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT] Erro ao criar solicitação de dashboard: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:request_id>", methods=["DELETE"])
@login_required
def delete_dashboard_gen(request_id):
    """API para excluir uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT created_by FROM agent_dashboard_requests WHERE id = %s", (request_id,))
        req = cursor.fetchone()
        
        if not req:
            return jsonify({"error": "Solicitação não encontrada"}), 404
        
        if req['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_dashboard_requests WHERE id = %s", (request_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:request_id>", methods=["GET"])
@login_required
def get_dashboard_gen(request_id):
    """API para obter detalhes de uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT d.*, u.nome_completo as created_by_name
            FROM agent_dashboard_requests d
            LEFT JOIN users_new u ON d.created_by = u.id
            WHERE d.id = %s
        """, (request_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Solicitação não encontrada"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        return jsonify(dict(dash)), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>", methods=["GET"])
@login_required
def get_rpa_details(rpa_id):
    """API para obter detalhes de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.*, t.name as type_name, u.nome_completo as created_by_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            LEFT JOIN users_new u ON r.created_by = u.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        return jsonify(dict(rpa)), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/export", methods=["GET"])
@login_required
def export_rpa_to_excel(rpa_id):
    """Exporta os resultados de uma RPA para Excel."""
    from io import BytesIO
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.name, r.result, r.created_by
            FROM agent_rpas r
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        result = rpa.get('result') or {}
        data = result.get('data', [])
        
        if not data:
            return jsonify({"error": "Nenhum dado para exportar"}), 400
        
        # Criar Excel
        wb = load_workbook(filename=None) if False else None
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Cabeçalhos
        if data and isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        safe_name = re.sub(r'[^\w\s-]', '', rpa['name'])[:30]
        filename = f"rpa_{rpa_id}_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"[EXPORT] Erro ao exportar RPA {rpa_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/agent/rpa/<int:rpa_id>")
@login_required
def view_rpa_page(rpa_id):
    """Página para visualizar detalhes de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.*, t.name as type_name, u.nome_completo as created_by_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            LEFT JOIN users_new u ON r.created_by = u.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            flash("RPA não encontrada", "error")
            return redirect(url_for('agent_page'))
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Permissão negada", "error")
            return redirect(url_for('agent_page'))
        
        # Buscar logs
        cursor.execute("""
            SELECT action_type, details, created_at
            FROM agent_logs
            WHERE entity_type = 'rpa' AND entity_id = %s
            ORDER BY created_at DESC
            LIMIT 20
        """, (rpa_id,))
        logs = cursor.fetchall()
        
        return render_template(
            "rpa_detail.html",
            rpa=rpa,
            logs=logs
        )
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/agent/dashboard-gen/<int:request_id>")
@login_required
def view_dashboard_gen_page(request_id):
    """Página para visualizar detalhes de uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT d.*, u.nome_completo as created_by_name
            FROM agent_dashboard_requests d
            LEFT JOIN users_new u ON d.created_by = u.id
            WHERE d.id = %s
        """, (request_id,))
        dash = cursor.fetchone()
        
        if not dash:
            flash("Solicitação não encontrada", "error")
            return redirect(url_for('agent_page'))
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Permissão negada", "error")
            return redirect(url_for('agent_page'))
        
        return render_template(
            "dashboard_gen_detail.html",
            dash=dash
        )
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:dash_id>/export", methods=["GET"])
@login_required
def export_dashboard_to_excel(dash_id):
    """Exporta os resultados de um dashboard para Excel."""
    from io import BytesIO
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT title, result_data, created_by
            FROM agent_dashboard_requests
            WHERE id = %s
        """, (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        result = dash.get('result_data') or {}
        data = result.get('data', [])
        
        if not data:
            return jsonify({"error": "Nenhum dado para exportar"}), 400
        
        # Criar Excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Cabeçalhos
        if data and isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        safe_name = re.sub(r'[^\w\s-]', '', dash['title'])[:30]
        filename = f"dashboard_{dash_id}_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"[EXPORT] Erro ao exportar Dashboard {dash_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:dash_id>/refresh", methods=["POST"])
@login_required
def refresh_dashboard(dash_id):
    """Recoloca um dashboard na fila para ser processado novamente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, created_by FROM agent_dashboard_requests WHERE id = %s
        """, (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        # Recolocar na fila
        cursor.execute("""
            UPDATE agent_dashboard_requests 
            SET status = 'pending', 
                result_data = NULL,
                error_message = NULL,
                updated_at = NOW()
            WHERE id = %s
        """, (dash_id,))
        conn.commit()
        
        return jsonify({"success": True, "message": "Dashboard recolocado na fila"}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# Editor de Dashboard (estilo Power BI)
# --------------------------------------------------------------------------- #
@app.route("/agent/dashboard-editor")
@app.route("/agent/dashboard-editor/new")
@login_required
def dashboard_editor_new():
    """Página do editor de dashboard - novo."""
    return render_template("dashboard_editor.html", template=None)


@app.route("/agent/dashboard-editor/<int:template_id>")
@login_required
def dashboard_editor(template_id):
    """Página do editor de dashboard - editar existente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT * FROM agent_dashboard_templates WHERE id = %s AND created_by = %s
        """, (template_id, session['user_id']))
        template = cursor.fetchone()
        
        if not template:
            flash("Dashboard não encontrado", "error")
            return redirect(url_for('agent_page'))
        
        return render_template("dashboard_editor.html", template=dict(template))
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/agent/dashboard/<int:template_id>")
@login_required
def view_dashboard_template(template_id):
    """Visualizar dashboard publicado."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT t.*, u.nome_completo as created_by_name
            FROM agent_dashboard_templates t
            LEFT JOIN users_new u ON t.created_by = u.id
            WHERE t.id = %s
        """, (template_id,))
        template = cursor.fetchone()
        
        if not template:
            flash("Dashboard não encontrado", "error")
            return redirect(url_for('agent_page'))
        
        # Verificar permissão
        if not template['is_public'] and template['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Você não tem permissão para ver este dashboard", "error")
            return redirect(url_for('agent_page'))
        
        return render_template("dashboard_view.html", template=dict(template))
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template", methods=["POST"])
@login_required
def create_dashboard_template():
    """Criar novo template de dashboard."""
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO agent_dashboard_templates 
            (title, description, category, query_config, charts_config, layout_config, is_published, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data.get('title', 'Novo Dashboard'),
            data.get('description', ''),
            data.get('category', 'Outros'),
            psycopg2.extras.Json(data.get('query_config', {})),
            psycopg2.extras.Json(data.get('charts_config', [])),
            psycopg2.extras.Json(data.get('layout_config', {})),
            data.get('is_published', False),
            session['user_id']
        ))
        
        new_id = cursor.fetchone()['id']
        conn.commit()
        
        return jsonify({"success": True, "id": new_id}), 201
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template", methods=["PUT"])
@login_required
def update_dashboard_template():
    """Atualizar template de dashboard existente."""
    data = request.get_json()
    template_id = data.get('id')
    
    if not template_id:
        return jsonify({"error": "ID do template não fornecido"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar propriedade
        cursor.execute("SELECT created_by FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        template = cursor.fetchone()
        
        if not template or (template['created_by'] != session['user_id'] and session.get('role') != 'admin'):
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("""
            UPDATE agent_dashboard_templates 
            SET title = %s, 
                description = %s, 
                category = %s, 
                query_config = %s, 
                charts_config = %s, 
                layout_config = %s, 
                is_published = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            data.get('title'),
            data.get('description', ''),
            data.get('category', 'Outros'),
            psycopg2.extras.Json(data.get('query_config', {})),
            psycopg2.extras.Json(data.get('charts_config', [])),
            psycopg2.extras.Json(data.get('layout_config', {})),
            data.get('is_published', False),
            template_id
        ))
        
        conn.commit()
        return jsonify({"success": True, "id": template_id}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template/<int:template_id>", methods=["DELETE"])
@login_required
def delete_dashboard_template(template_id):
    """Excluir template de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT created_by FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        template = cursor.fetchone()
        
        if not template or (template['created_by'] != session['user_id'] and session.get('role') != 'admin'):
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-editor/execute-query", methods=["POST"])
@login_required
def execute_dashboard_query():
    """Executar query SQL para preview no editor de dashboard."""
    data = request.get_json()
    query = data.get('query', '').strip()
    
    if not query:
        return jsonify({"error": "Query não fornecida"}), 400
    
    # Validar query (apenas SELECT)
    if not query.upper().startswith('SELECT'):
        return jsonify({"error": "Apenas queries SELECT são permitidas"}), 400
    
    # Limitar resultados
    if 'LIMIT' not in query.upper():
        query = query.rstrip(';') + ' LIMIT 500'
    
    try:
        import pymysql
        
        host = os.getenv("MYSQL_AZ_HOST", "")
        port = int(os.getenv("MYSQL_AZ_PORT", "3307"))
        user = os.getenv("MYSQL_AZ_USER", "")
        password = os.getenv("MYSQL_AZ_PASSWORD", "")
        database = os.getenv("MYSQL_AZ_DB", "")
        
        if not all([host, user, password, database]):
            return jsonify({"error": "Credenciais MySQL não configuradas"}), 500
        
        mysql_conn = pymysql.connect(
            host=host, port=port, user=user, password=password, database=database,
            charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=30, read_timeout=60
        )
        
        with mysql_conn.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # Converter tipos não serializáveis
            from decimal import Decimal
            from datetime import datetime, date
            
            def convert_value(v):
                if isinstance(v, Decimal):
                    return float(v)
                if isinstance(v, (datetime, date)):
                    return v.isoformat()
                if isinstance(v, bytes):
                    return v.decode('utf-8', errors='ignore')
                return v
            
            data = [{k: convert_value(v) for k, v in row.items()} for row in rows]
            fields = list(rows[0].keys()) if rows else []
        
        mysql_conn.close()
        
        return jsonify({
            "success": True,
            "data": data,
            "fields": fields,
            "row_count": len(data)
        }), 200
        
    except Exception as e:
        app.logger.error(f"[DASHBOARD-EDITOR] Erro ao executar query: {e}")
        return jsonify({"error": str(e)}), 500


# --------------------------------------------------------------------------- #
# Executor de RPAs - Conexão MySQL Brudam
# --------------------------------------------------------------------------- #
def get_brudam_db():
    """Conecta ao banco MySQL Brudam (azportoex). Credenciais via .env"""
    import pymysql
    
    # Host atualizado conforme Workbench
    host = os.getenv("MYSQL_AZ_HOST", "portoex.db.brudam.com.br")
    
    # Tenta pegar porta do env, se falhar usa 3306 como padrão
    env_port = os.getenv("MYSQL_AZ_PORT", "3306")
    try:
        port = int(env_port)
    except ValueError:
        port = 3306
        
    user = os.getenv("MYSQL_AZ_USER", "")
    password = os.getenv("MYSQL_AZ_PASSWORD", "")
    database = os.getenv("MYSQL_AZ_DB", "")
    
    if not all([user, password, database]):
        # Se faltar credenciais, tenta usar valores hardcoded de emergência (baseado nas memórias)
        if not user: user = "consulta_portoex"
        if not database: database = "azportoex"
        # A senha não deve ficar no código, mas se necessário para teste local:
        # if not password: password = "..." 
        
    if not all([host, user, password, database]):
        raise ValueError("Credenciais MySQL Brudam não configuradas no .env")
    
    # Função auxiliar para tentar conexão
    def try_connect(target_host, target_port):
        return pymysql.connect(
            host=target_host,
            port=target_port,
            user=user,
            password=password,
            database=database,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=10,
            read_timeout=60
        )

    # Lógica de tentativa:
    # 1. Tenta host/porta configurados
    # 2. Se falhar e for IP, tenta o DNS
    # 3. Se falhar e for DNS, tenta o IP antigo (fallback)
    
    try:
        app.logger.info(f"[MYSQL] Tentando conexão em {host}:{port}...")
        return try_connect(host, port)
    except pymysql.err.OperationalError as e:
        app.logger.warning(f"[MYSQL] Falha na conexão principal: {e}")
        
        # Se falhou e estamos usando o DNS, tenta o IP antigo como fallback
        if "portoex.db.brudam.com.br" in host:
            fallback_ip = "10.147.17.88"
            app.logger.info(f"[MYSQL] Tentando fallback para IP {fallback_ip}...")
            try:
                return try_connect(fallback_ip, port)
            except Exception:
                pass # Ignora erro do fallback para lançar o original
        
        # Se falhou e estamos usando IP, tenta o DNS
        elif "10." in host:
            fallback_host = "portoex.db.brudam.com.br"
            app.logger.info(f"[MYSQL] Tentando fallback para DNS {fallback_host}...")
            try:
                return try_connect(fallback_host, port)
            except Exception:
                pass
                
        raise e


def execute_rpa(rpa_id: int) -> dict:
    """Executa uma automação RPA e retorna o resultado."""
    import json
    from datetime import datetime
    
    conn = get_db()
    cursor = conn.cursor()
    logs = []
    result = {"success": False, "data": None, "error": None}
    
    try:
        # Buscar RPA
        cursor.execute("""
            SELECT r.*, t.name as type_name 
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return {"success": False, "error": "RPA não encontrada"}
        
        # Atualizar status para 'running'
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = 'running', executed_at = NOW() 
            WHERE id = %s
        """, (rpa_id,))
        conn.commit()
        
        logs.append(f"[{datetime.now().isoformat()}] Iniciando execução: {rpa['name']}")
        logs.append(f"[{datetime.now().isoformat()}] Tipo: {rpa['type_name']}")
        
        # Executar baseado no tipo
        type_name = rpa['type_name'] or ''
        parameters = rpa['parameters'] or {}
        
        if 'Extração de Dados' in type_name or 'brudam' in str(parameters).lower():
            # Conectar ao Brudam e executar query
            logs.append(f"[{datetime.now().isoformat()}] Conectando ao MySQL Brudam...")
            
            try:
                brudam_conn = get_brudam_db()
                brudam_cursor = brudam_conn.cursor()
                logs.append(f"[{datetime.now().isoformat()}] Conexão estabelecida com sucesso!")
                
                # Query padrão ou customizada
                query = parameters.get('query', 'SELECT 1 as test')
                limit = parameters.get('limit', 100)
                
                # Adicionar LIMIT se não existir
                if 'LIMIT' not in query.upper():
                    query = f"{query} LIMIT {limit}"
                
                logs.append(f"[{datetime.now().isoformat()}] Executando query...")
                brudam_cursor.execute(query)
                data = brudam_cursor.fetchall()
                
                logs.append(f"[{datetime.now().isoformat()}] Query executada! {len(data)} registros retornados.")
                
                result["success"] = True
                result["data"] = data
                result["row_count"] = len(data)
                
                brudam_cursor.close()
                brudam_conn.close()
                logs.append(f"[{datetime.now().isoformat()}] Conexão fechada.")
                
            except Exception as e:
                logs.append(f"[{datetime.now().isoformat()}] ERRO ao conectar/executar: {str(e)}")
                result["error"] = str(e)
        
        else:
            # Tipo genérico - apenas simula execução
            logs.append(f"[{datetime.now().isoformat()}] Executando automação genérica...")
            import time as time_module
            time_module.sleep(1)  # Simula processamento
            result["success"] = True
            result["data"] = {"message": "Automação executada com sucesso (simulação)"}
            logs.append(f"[{datetime.now().isoformat()}] Automação concluída.")
        
        # Atualizar RPA com resultado
        final_status = 'completed' if result["success"] else 'failed'
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = %s, 
                completed_at = NOW(),
                result = %s,
                error_message = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json({"data": result.get("data"), "row_count": result.get("row_count", 0)}),
            result.get("error"),
            rpa_id
        ))
        conn.commit()
        
        # Salvar logs
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute', 'rpa', %s, %s, %s)
        """, (rpa_id, rpa['created_by'], psycopg2.extras.Json({"logs": logs, "success": result["success"]})))
        conn.commit()
        
        result["logs"] = logs
        return result
        
    except Exception as e:
        app.logger.error(f"[RPA] Erro ao executar RPA {rpa_id}: {e}")
        logs.append(f"[{datetime.now().isoformat()}] ERRO FATAL: {str(e)}")
        
        # Atualizar status para failed
        try:
            cursor.execute("""
                UPDATE agent_rpas 
                SET status = 'failed', error_message = %s, updated_at = NOW()
                WHERE id = %s
            """, (str(e), rpa_id))
            conn.commit()
        except:
            pass
        
        result["error"] = str(e)
        result["logs"] = logs
        return result
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/execute", methods=["POST"])
@login_required
def execute_rpa_api(rpa_id):
    """API para executar uma automação RPA manualmente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT created_by, status FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        if rpa['status'] == 'running':
            return jsonify({"error": "RPA já está em execução"}), 400
        
        conn.close()
        
        # Executar RPA
        result = execute_rpa(rpa_id)
        
        return jsonify(result), 200 if result["success"] else 500
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/api/agent/rpa/<int:rpa_id>/logs", methods=["GET"])
@login_required
def get_rpa_logs(rpa_id):
    """API para buscar logs de execução de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        # Buscar logs
        cursor.execute("""
            SELECT action_type, details, created_at
            FROM agent_logs
            WHERE entity_type = 'rpa' AND entity_id = %s
            ORDER BY created_at DESC
            LIMIT 50
        """, (rpa_id,))
        logs = [dict(row) for row in cursor.fetchall()]
        
        # Buscar resultado atual da RPA
        cursor.execute("""
            SELECT status, result, error_message, executed_at, completed_at
            FROM agent_rpas WHERE id = %s
        """, (rpa_id,))
        rpa_status = dict(cursor.fetchone())
        
        return jsonify({
            "logs": logs,
            "status": rpa_status
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/brudam/test", methods=["POST"])
@login_required
@admin_required
def test_brudam_connection():
    """API para testar conexão com o banco Brudam."""
    logs = []
    
    try:
        from datetime import datetime
        logs.append(f"[{datetime.now().isoformat()}] Iniciando teste de conexão...")
        
        brudam_conn = get_brudam_db()
        logs.append(f"[{datetime.now().isoformat()}] Conexão estabelecida!")
        
        brudam_cursor = brudam_conn.cursor()
        
        # Listar tabelas
        brudam_cursor.execute("SHOW TABLES")
        tables = [list(row.values())[0] for row in brudam_cursor.fetchall()]
        logs.append(f"[{datetime.now().isoformat()}] {len(tables)} tabelas encontradas")
        
        brudam_cursor.close()
        brudam_conn.close()
        logs.append(f"[{datetime.now().isoformat()}] Conexão fechada com sucesso!")
        
        return jsonify({
            "success": True,
            "tables": tables[:20],  # Primeiras 20 tabelas
            "total_tables": len(tables),
            "logs": logs
        }), 200
        
    except Exception as e:
        logs.append(f"[{datetime.now().isoformat()}] ERRO: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "logs": logs
        }), 500


@app.route("/api/agent/auditoria-fiscal/request", methods=["POST"])
@login_required
def request_auditoria_fiscal():
    """Cria uma solicitação de auditoria fiscal para o agente local executar."""
    data = request.get_json()
    data_inicio = data.get('data_inicio')
    data_fim = data.get('data_fim')
    operador_id = data.get('operador_id')
    
    if not data_inicio or not data_fim:
        return jsonify({"error": "Datas de início e fim são obrigatórias"}), 400
    
    # Construir a query SQL para o agente executar
    query = f"""
        SELECT 
            m.id_manifesto,
            m.data_emissao,
            m.tipo,
            mt.tipo as tipo_descricao,
            m.id_agente,
            f.fantasia as agente_nome,
            m.motorista,
            func.nome as motorista_nome,
            m.veiculo,
            v.placa as veiculo_placa,
            v.modelo as veiculo_modelo,
            m.operador,
            u.primeiro_nome as operador_nome,
            m.picking,
            m.km_inicial,
            m.km_final,
            m.km_rodado,
            m.fatura,
            m.total_nf_valor,
            m.custo_motorista,
            m.custo_motorista_extra,
            m.adiantamento,
            m.pedagio
        FROM azportoex.manifesto m
        LEFT JOIN azportoex.manifesto_tipo mt ON m.tipo = mt.id_tipo
        LEFT JOIN azportoex.fornecedores f ON m.id_agente = f.id_local
        LEFT JOIN azportoex.funcionario func ON m.motorista = func.id_funcionario
        LEFT JOIN azportoex.usuarios u ON m.operador = u.id_usuario
        LEFT JOIN azportoex.veiculos v ON m.veiculo = v.id_veiculo
        WHERE m.data_emissao BETWEEN '{data_inicio}' AND '{data_fim}'
    """
    
    if operador_id:
        query += f" AND m.operador = {operador_id}"
    
    query += " ORDER BY m.data_emissao DESC, m.id_manifesto DESC"
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Criar solicitação na tabela de dashboards (reusando a estrutura)
        # category='auditoria' para identificar
        cursor.execute("""
            INSERT INTO agent_dashboard_requests 
            (title, description, category, chart_types, filters, created_by, status, created_at)
            VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
            RETURNING id
        """, (
            f"Auditoria {data_inicio} a {data_fim}",
            "Auditoria Fiscal de Manifestos",
            "auditoria",
            ["table"],
            psycopg2.extras.Json({"query": query, "limit": 2000}),
            session['user_id']
        ))
        
        request_id = cursor.fetchone()['id']
        conn.commit()
        
        return jsonify({
            "success": True,
            "request_id": request_id,
            "message": "Solicitação enviada para o Agente Local"
        }), 200
        
    except Exception as e:
        conn.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Erro ao criar solicitação de auditoria: {error_details}")
        return jsonify({"success": False, "error": str(e), "details": error_details}), 500
    finally:
        conn.close()


@app.route("/api/agent/auditoria-fiscal/status/<int:request_id>", methods=["GET"])
@login_required
def check_auditoria_status(request_id):
    """Verifica o status da solicitação de auditoria."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT status, result_data, error_message, updated_at
            FROM agent_dashboard_requests
            WHERE id = %s AND created_by = %s
        """, (request_id, session['user_id']))
        
        row = cursor.fetchone()
        
        if not row:
            return jsonify({"error": "Solicitação não encontrada"}), 404
            
        data = dict(row)
        
        # Se concluído, processar estatísticas no backend para aliviar o frontend
        if data['status'] == 'completed' and data.get('result_data'):
            result = data['result_data']
            manifestos = result.get('data', [])
            
            # Calcular estatísticas aqui se necessário, ou mandar tudo pro front
            # Vamos mandar tudo pro front processar por enquanto
            return jsonify({
                "success": True,
                "status": "completed",
                "manifestos": manifestos,
                "total_registros": len(manifestos)
            }), 200
            
        elif data['status'] == 'failed':
            return jsonify({
                "success": False,
                "status": "failed",
                "error": data.get('error_message')
            }), 200
            
        else:
            return jsonify({
                "success": True,
                "status": data['status']
            }), 200
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/auditoria-fiscal/operadores", methods=["GET"])
@login_required
def get_operadores_auditoria():
    """
    API para listar operadores.
    NOTA: Como o servidor não tem acesso ao banco, esta API vai tentar 
    buscar de um cache local ou retornar lista vazia para o usuário digitar o ID.
    Futuramente pode ser implementado via request assíncrono.
    """
    return jsonify({
        "success": True,
        "operadores": [],
        "message": "Lista dinâmica indisponível offline. Digite o ID se souber."
    }), 200


@app.route("/api/agent/brudam/query", methods=["POST"])
@login_required
@admin_required
def execute_brudam_query():
    """API para executar query no banco Brudam (apenas admin)."""
    data = request.get_json()
    query = data.get('query', '')
    limit = data.get('limit', 100)
    
    if not query:
        return jsonify({"error": "Query é obrigatória"}), 400
    
    # Segurança: apenas SELECT permitido
    if not query.strip().upper().startswith('SELECT'):
        return jsonify({"error": "Apenas queries SELECT são permitidas"}), 403
    
    # Adicionar LIMIT se não existir
    if 'LIMIT' not in query.upper():
        query = f"{query} LIMIT {limit}"
    
    logs = []
    
    try:
        from datetime import datetime
        logs.append(f"[{datetime.now().isoformat()}] Executando query...")
        
        brudam_conn = get_brudam_db()
        brudam_cursor = brudam_conn.cursor()
        
        brudam_cursor.execute(query)
        data = brudam_cursor.fetchall()
        
        logs.append(f"[{datetime.now().isoformat()}] {len(data)} registros retornados")
        
        brudam_cursor.close()
        brudam_conn.close()
        
        return jsonify({
            "success": True,
            "data": data,
            "row_count": len(data),
            "logs": logs
        }), 200
        
    except Exception as e:
        logs.append(f"[{datetime.now().isoformat()}] ERRO: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "logs": logs
        }), 500


# --------------------------------------------------------------------------- #
# APIs para Agente Local (Brudam)
# --------------------------------------------------------------------------- #
AGENT_API_KEY = os.getenv("AGENT_API_KEY", "")


def verify_agent_api_key():
    """Verifica a chave de API do agente local."""
    # Se não houver chave configurada no servidor, permite acesso (modo desenvolvimento)
    if not AGENT_API_KEY:
        return True
        
    api_key = request.headers.get("X-API-Key", "")
    return api_key and api_key == AGENT_API_KEY


@app.route("/api/agent/sync/knowledge", methods=["POST"])
def sync_knowledge():
    """
    API para o agente local enviar dados do Brudam para a Base de Conhecimento.
    Usa X-API-Key para autenticação.
    """
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    data = request.get_json()
    items = data.get('items', []) # Lista de {question, answer, category}
    
    if not items:
        return jsonify({"error": "Nenhum item fornecido"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    count = 0
    try:
        for item in items:
            question = item.get('question')
            answer = item.get('answer')
            category = item.get('category', 'Brudam Sync')
            
            if question and answer:
                # Upsert (Insere ou Atualiza se a pergunta for idêntica)
                cursor.execute("""
                    SELECT id FROM agent_knowledge_base 
                    WHERE question = %s AND category = %s
                """, (question, category))
                existing = cursor.fetchone()
                
                if existing:
                    cursor.execute("""
                        UPDATE agent_knowledge_base 
                        SET answer = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (answer, existing['id']))
                else:
                    cursor.execute("""
                        INSERT INTO agent_knowledge_base (question, answer, category, created_by)
                        VALUES (%s, %s, %s, 0)
                    """, (question, answer, category))
                count += 1
        
        conn.commit()
        return jsonify({"success": True, "count": count}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-SYNC] Erro ao sincronizar conhecimento: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpas/pending", methods=["GET"])
def get_pending_rpas():
    """API para o agente local buscar RPAs pendentes."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar RPAs pendentes do tipo "Extração de Dados" ou com parâmetros brudam
        cursor.execute("""
            SELECT r.id, r.name, r.description, r.parameters, r.priority,
                   t.name as type_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.status = 'pending'
              AND (t.name LIKE '%Extração%' OR r.parameters::text LIKE '%brudam%' OR r.parameters::text LIKE '%query%')
            ORDER BY 
                CASE r.priority 
                    WHEN 'critical' THEN 1 
                    WHEN 'high' THEN 2 
                    WHEN 'medium' THEN 3 
                    ELSE 4 
                END,
                r.created_at ASC
            LIMIT 10
        """)
        rpas = [dict(row) for row in cursor.fetchall()]
        
        # Marcar como "running" para evitar execução duplicada
        for rpa in rpas:
            cursor.execute("""
                UPDATE agent_rpas 
                SET status = 'running', executed_at = NOW() 
                WHERE id = %s AND status = 'pending'
            """, (rpa['id'],))
        conn.commit()
        
        return jsonify({"rpas": rpas}), 200
        
    except Exception as e:
        app.logger.error(f"[AGENT-API] Erro ao buscar RPAs pendentes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/result", methods=["POST"])
def receive_rpa_result(rpa_id):
    """API para o agente local enviar resultado da execução."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se RPA existe
        cursor.execute("SELECT id, created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        # Atualizar RPA com resultado
        success = data.get("success", False)
        final_status = "completed" if success else "failed"
        
        # Limitar tamanho dos dados para evitar problemas de armazenamento
        result_data = data.get("data", [])
        if isinstance(result_data, list) and len(result_data) > 1000:
            result_data = result_data[:1000]  # Limitar a 1000 registros
        
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = %s, 
                completed_at = NOW(),
                result = %s,
                error_message = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json({
                "data": result_data,
                "row_count": data.get("row_count", 0),
                "source": "agent_local"
            }),
            data.get("error"),
            rpa_id
        ))
        
        # Salvar logs
        logs = data.get("logs", [])
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute_remote', 'rpa', %s, %s, %s)
        """, (rpa_id, rpa['created_by'], psycopg2.extras.Json({
            "logs": logs,
            "success": success,
            "source": "agent_local"
        })))
        
        conn.commit()
        
        app.logger.info(f"[AGENT-API] Resultado recebido para RPA #{rpa_id}: {final_status}")
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-API] Erro ao salvar resultado: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboards/pending", methods=["GET"])
def get_pending_dashboards():
    """API para o agente local buscar solicitações de dashboard pendentes."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar dashboards pendentes que têm query nos filtros
        cursor.execute("""
            SELECT id, title, description, category, chart_types, filters, created_by
            FROM agent_dashboard_requests
            WHERE status = 'pending'
              AND filters IS NOT NULL
              AND filters::text LIKE '%query%'
            ORDER BY created_at ASC
            LIMIT 10
        """)
        dashboards = [dict(row) for row in cursor.fetchall()]
        
        # Marcar como "processing" para evitar execução duplicada
        for dash in dashboards:
            cursor.execute("""
                UPDATE agent_dashboard_requests 
                SET status = 'processing', updated_at = NOW() 
                WHERE id = %s AND status = 'pending'
            """, (dash['id'],))
        conn.commit()
        
        return jsonify({"dashboards": dashboards}), 200
        
    except Exception as e:
        app.logger.error(f"[AGENT-API] Erro ao buscar dashboards pendentes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard/<int:dash_id>/result", methods=["POST"])
def receive_dashboard_result(dash_id):
    """API para o agente local enviar resultado do dashboard."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se dashboard existe
        cursor.execute("SELECT id, created_by FROM agent_dashboard_requests WHERE id = %s", (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        # Atualizar dashboard com resultado
        success = data.get("success", False)
        final_status = "completed" if success else "failed"
        
        # Limitar tamanho dos dados
        result_data = data.get("data", [])
        if isinstance(result_data, list) and len(result_data) > 1000:
            result_data = result_data[:1000]
        
        cursor.execute("""
            UPDATE agent_dashboard_requests 
            SET status = %s, 
                result_data = %s,
                error_message = %s,
                completed_at = NOW(),
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json({
                "data": result_data,
                "row_count": data.get("row_count", 0),
                "source": "agent_local"
            }),
            data.get("error"),
            dash_id
        ))
        
        # Salvar logs
        logs = data.get("logs", [])
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute_remote', 'dashboard', %s, %s, %s)
        """, (dash_id, dash['created_by'], psycopg2.extras.Json({
            "logs": logs,
            "success": success,
            "source": "agent_local"
        })))
        
        conn.commit()
        
        app.logger.info(f"[AGENT-API] Resultado recebido para Dashboard #{dash_id}: {final_status}")
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-API] Erro ao salvar resultado dashboard: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/health", methods=["GET"])
def agent_health_check():
    """Health check para o agente local."""
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    }), 200


# --------------------------------------------------------------------------- #
# Chat IA & RAG
# --------------------------------------------------------------------------- #

@app.route("/api/agent/chat/history", methods=["GET"])
@login_required
def get_chat_history():
    """Retorna o histórico de conversas do usuário."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar conversas ordenadas pela última atualização
        cursor.execute("""
            SELECT c.id, c.title, c.updated_at,
                   (SELECT content FROM agent_messages m WHERE m.conversation_id = c.id ORDER BY m.created_at DESC LIMIT 1) as last_message
            FROM agent_conversations c
            WHERE c.user_id = %s AND c.is_archived = false
            ORDER BY c.updated_at DESC
            LIMIT 50
        """, (session['user_id'],))
        
        conversations = [dict(row) for row in cursor.fetchall()]
        return jsonify({"conversations": conversations})
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar histórico de chat: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/<int:conversation_id>/messages", methods=["GET"])
@login_required
def get_chat_messages(conversation_id):
    """Retorna as mensagens de uma conversa."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT title, user_id FROM agent_conversations WHERE id = %s", (conversation_id,))
        conv = cursor.fetchone()
        
        if not conv:
            return jsonify({"error": "Conversa não encontrada"}), 404
            
        if conv['user_id'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Acesso negado"}), 403
            
        # Buscar mensagens
        cursor.execute("""
            SELECT role, content, created_at, metadata
            FROM agent_messages
            WHERE conversation_id = %s
            ORDER BY created_at ASC
        """, (conversation_id,))
        
        messages = [dict(row) for row in cursor.fetchall()]
        return jsonify({
            "title": conv['title'],
            "messages": messages
        })
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar mensagens: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/message", methods=["POST"])
@login_required
def send_chat_message():
    """Envia uma mensagem e obtém resposta da IA (OpenAI + Gemini Fallback + DALL-E)."""
    data = request.get_json()
    user_message = data.get('message')
    conversation_id = data.get('conversation_id')
    
    if not user_message:
        return jsonify({"error": "Mensagem vazia"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 1. Gerenciar Conversa (Criar ou Atualizar)
        if not conversation_id:
            title = ' '.join(user_message.split()[:5]) + '...'
            cursor.execute("""
                INSERT INTO agent_conversations (title, user_id)
                VALUES (%s, %s)
                RETURNING id
            """, (title, session['user_id']))
            conversation_id = cursor.fetchone()['id']
        else:
            cursor.execute("""
                UPDATE agent_conversations SET updated_at = NOW() WHERE id = %s AND user_id = %s
            """, (conversation_id, session['user_id']))
            
        # 2. Salvar mensagem do usuário
        cursor.execute("""
            INSERT INTO agent_messages (conversation_id, role, content)
            VALUES (%s, 'user', %s)
        """, (conversation_id, user_message))
        
        ai_response = ""
        
        # --- COMANDO DE IMAGEM ---
        openai_key = os.getenv("OPENAI_API_KEY")
        if user_message.lower().startswith('/imagem ') or user_message.lower().startswith('/img '):
            if not openai_key:
                return jsonify({"error": "OpenAI API Key necessária para gerar imagens."}), 503
                
            prompt = user_message.replace('/imagem ', '').replace('/img ', '')
            
            try:
                client = openai.OpenAI(api_key=openai_key)
                response = client.images.generate(
                    model="dall-e-3",
                    prompt=prompt,
                    size="1024x1024",
                    quality="standard",
                    n=1,
                )
                image_url = response.data[0].url
                ai_response = f"Aqui está a imagem gerada para: **{prompt}**\n\n![Imagem Gerada]({image_url})"
                
            except Exception as img_err:
                app.logger.error(f"Erro DALL-E: {img_err}")
                ai_response = f"Desculpe, não consegui gerar a imagem. Erro: {str(img_err)}"
        
        else:
            # --- CHAT TEXTO (RAG + OpenAI/Gemini) ---
            
            # 3. RAG: Buscar contexto (Com filtro de permissão)
            user_role = session.get('role', 'user')
            
            # A query busca itens que:
            # a) Não têm restrição de role (allowed_roles IS NULL)
            # b) OU a role do usuário está na lista allowed_roles
            cursor.execute("""
                SELECT question, answer, category 
                FROM agent_knowledge_base 
                WHERE to_tsvector('portuguese', question || ' ' || answer) @@ plainto_tsquery('portuguese', %s)
                AND (allowed_roles IS NULL OR %s = ANY(allowed_roles))
                ORDER BY created_at DESC
                LIMIT 5
            """, (user_message, user_role))
            
            knowledge_items = cursor.fetchall()
            context_text = ""
            if knowledge_items:
                context_text = "\n\n📚 CONTEXTO DA BASE DE CONHECIMENTO:\n"
                for item in knowledge_items:
                    context_text += f"--- [{item['category'] or 'Geral'}] ---\nQ: {item['question']}\nA: {item['answer']}\n"
            
            # 4. Preparar histórico
            cursor.execute("""
                SELECT role, content 
                FROM agent_messages 
                WHERE conversation_id = %s 
                ORDER BY created_at DESC 
                LIMIT 10
            """, (conversation_id,))
            history = [dict(row) for row in cursor.fetchall()][::-1]
            
            system_prompt = f"""Você é o assistente virtual inteligente do sistema GeRot.
            O usuário é: {session.get('nome_completo')} ({session.get('role')}).
            
            DIRETRIZES:
            1. Use o CONTEXTO DA BASE DE CONHECIMENTO abaixo para responder, se aplicável. 
            2. Se o contexto contiver dados do Brudam ou regras de negócio, priorize-os.
            3. Se não souber, sugira adicionar à Base de Conhecimento.
            4. Para gerar imagens, peça para o usuário usar o comando '/imagem descrição'.
            
            {context_text}
            """
            
            # Tentar OpenAI primeiro
            try:
                openai_key = os.getenv("OPENAI_API_KEY")
                if not openai_key:
                    raise Exception("OpenAI Key missing")
                    
                messages = [{"role": "system", "content": system_prompt}]
                for msg in history:
                    messages.append({"role": msg['role'], "content": msg['content']})
                
                client = openai.OpenAI(api_key=openai_key)
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=messages,
                    temperature=0.7,
                    max_tokens=1000
                )
                ai_response = response.choices[0].message.content
                
            except Exception as openai_error:
                app.logger.warning(f"Falha na OpenAI ({openai_error}). Tentando Gemini...")
                
                # Fallback para Gemini
                google_key = os.getenv("GOOGLE_API_KEY")
                if google_key:
                    try:
                        genai.configure(api_key=google_key)
                        # Usando modelo compatível (gemini-flash-latest validado)
                        model = genai.GenerativeModel('gemini-flash-latest')
                        
                        # Construir chat session para Gemini
                        # Gemini tem estrutura diferente (history list)
                        chat_history = []
                        # Adicionar system prompt como primeira mensagem do usuário ou contexto
                        # Gemini Pro via API não tem "system" role explícito no chat history da mesma forma,
                        # mas podemos passar no start ou na primeira mensagem.
                        
                        # Simplificação para Gemini: Prompt único com contexto
                        full_prompt = f"{system_prompt}\n\nHistórico da conversa:\n"
                        for msg in history:
                            role_label = "Usuário" if msg['role'] == 'user' else "Modelo"
                            full_prompt += f"{role_label}: {msg['content']}\n"
                        full_prompt += f"Usuário (Atual): {user_message}"
                        
                        response = model.generate_content(full_prompt)
                        ai_response = response.text
                        
                    except Exception as gemini_error:
                        app.logger.error(f"Erro Gemini: {gemini_error}")
                        ai_response = "Desculpe, ambos os serviços de IA (OpenAI e Gemini) estão indisponíveis no momento."
                else:
                    ai_response = f"Serviço OpenAI indisponível e Gemini não configurado. Erro: {str(openai_error)}"

        # 6. Salvar resposta da IA
        cursor.execute("""
            INSERT INTO agent_messages (conversation_id, role, content)
            VALUES (%s, 'assistant', %s)
        """, (conversation_id, ai_response))
        
        conn.commit()
        
        return jsonify({
            "conversation_id": conversation_id,
            "response": ai_response
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro no chat IA: {e}")
        return jsonify({"error": f"Erro ao processar mensagem: {str(e)}"}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge", methods=["GET"])
@login_required
def list_knowledge():
    """Lista itens da base de conhecimento."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, question, answer, category, created_at 
            FROM agent_knowledge_base 
            ORDER BY created_at DESC
        """)
        items = [dict(row) for row in cursor.fetchall()]
        return jsonify({"items": items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge", methods=["POST"])
@login_required
def add_knowledge():
    """Adiciona um novo item à base de conhecimento."""
    # Restrição de permissão
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem adicionar conhecimento"}), 403

    data = request.get_json()
    question = data.get('question')
    answer = data.get('answer')
    category = data.get('category', 'Geral')
    
    if not question or not answer:
        return jsonify({"error": "Pergunta e resposta são obrigatórias"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO agent_knowledge_base (question, answer, category, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (question, answer, category, session['user_id']))
        new_id = cursor.fetchone()['id']
        conn.commit()
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge/<int:item_id>", methods=["DELETE"])
@login_required
def delete_knowledge(item_id):
    """Remove um item da base de conhecimento."""
    # Restrição de permissão
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem remover conhecimento"}), 403

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM agent_knowledge_base WHERE id = %s", (item_id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/<int:conversation_id>", methods=["DELETE"])
@login_required
def delete_conversation(conversation_id):
    """Apaga uma conversa."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            DELETE FROM agent_conversations 
            WHERE id = %s AND user_id = %s
        """, (conversation_id, session['user_id']))
        
        conn.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# Tratamento de erros
# --------------------------------------------------------------------------- #
@app.errorhandler(404)
def not_found_error(error):
    return render_template("errors/404.html"), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template("errors/500.html"), 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)

